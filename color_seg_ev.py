
'''
Name: color_segmentation.py

Version: 1.0

Summary: K-means color clustering based segmentation. This is achieved 
         by converting the source image to a desired color space and 
         running K-means clustering on only the desired channels, 
         with the pixels being grouped into a desired number
    of clusters. 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2019-09-29

USAGE:

python3 color_seg_ev.py -p ~/example/plant_test/ -ft jpg -c 0 -min 100  -max 1500

python3 color_seg_ev.py -p ~/example/plant_test/ -ft jpg -nr 4 -nc 6

'''

# import the necessary packages
import os
import glob
import argparse
from sklearn.cluster import KMeans

from skimage.feature import peak_local_max
from skimage.morphology import watershed, medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.segmentation import clear_border

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage

import math

import numpy as np
import argparse
import cv2

import openpyxl
import csv

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import warnings
warnings.filterwarnings("ignore")

import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

MBFACTOR = float(1<<20)

from scipy.spatial import distance as dist
from collections import OrderedDict



class ColorLabeler:
    def __init__(self):
        # initialize the colors dictionary, containing the color
        # name as the key and the RGB tuple as the value
        colors = OrderedDict({
            "red": (255, 0, 0),
            "green": (0, 255, 0),
            "blue": (0, 0, 255)})
        # allocate memory for the L*a*b* image, then initialize
        # the color names list
        self.lab = np.zeros((len(colors), 1, 3), dtype="uint8")
        self.colorNames = []
        # loop over the colors dictionary
        for (i, (name, rgb)) in enumerate(colors.items()):
            # update the L*a*b* array and the color names list
            self.lab[i] = rgb
            self.colorNames.append(name)
        # convert the L*a*b* array from the RGB color space
        # to L*a*b*
        self.lab = cv2.cvtColor(self.lab, cv2.COLOR_RGB2LAB)


    def label(self, image, c):
        # construct a mask for the contour, then compute the
        # average L*a*b* value for the masked region
        mask = np.zeros(image.shape[:2], dtype="uint8")
        cv2.drawContours(mask, [c], -1, 255, -1)
        mask = cv2.erode(mask, None, iterations=2)
        mean = cv2.mean(image, mask=mask)[:3]
        # initialize the minimum distance found thus far
        minDist = (np.inf, None)
        # loop over the known L*a*b* color values
        for (i, row) in enumerate(self.lab):
            # compute the distance between the current L*a*b*
            # color value and the mean of the image
            d = dist.euclidean(row[0], mean)
            # if the distance is smaller than the current distance,
            # then update the bookkeeping variable
            if d < minDist[0]:
                minDist = (d, i)
        # return the name of the color with the smallest distance
        return self.colorNames[minDist[1]]



class clockwise_angle_and_distance():
    

    '''
    A class to tell if point is clockwise from origin or not.
    This helps if one wants to use sorted() on a list of points.
    Parameters
    ----------
    point : ndarray or list, like [x, y]. The point "to where" we g0
    self.origin : ndarray or list, like [x, y]. The center around which we go
    refvec : ndarray or list, like [x, y]. The direction of reference
    use: 
        instantiate with an origin, then call the instance during sort
    reference: 
    https://stackoverflow.com/questions/41855695/sorting-list-of-two-dimensional-coordinates-by-clockwise-angle-using-python
    Returns
    -------
    angle
    
    distance
    
    '''
    def __init__(self, origin):
        self.origin = origin

    def __call__(self, point, refvec = [0, 1]):
        if self.origin is None:
            raise NameError("clockwise sorting needs an origin. Please set origin.")
        # Vector between point and the origin: v = p - o
        vector = [point[0]-self.origin[0], point[1]-self.origin[1]]
        # Length of vector: ||v||
        lenvector = np.linalg.norm(vector[0] - vector[1])
        # If length is zero there is no angle
        if lenvector == 0:
            return -pi, 0
        # Normalize vector: v/||v||
        normalized = [vector[0]/lenvector, vector[1]/lenvector]
        dotprod  = normalized[0]*refvec[0] + normalized[1]*refvec[1] # x1*x2 + y1*y2
        diffprod = refvec[1]*normalized[0] - refvec[0]*normalized[1] # x1*y2 - y1*x2
        angle = math.atan2(diffprod, dotprod)
        # Negative angles represent counter-clockwise angles so we need to 
        # subtract them from 2*pi (360 degrees)
        if angle < 0:
            return 2*math.pi+angle, lenvector
        # I return first the angle because that's the primary sorting criterium
        # but if two vectors have the same angle then the shorter distance 
        # should come first.
        return angle, lenvector



# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False


def sort_contours(contours):

    # initialize the reverse flag and sort index
    reverse = False
    i = 0
  
    # handle if we need to sort in reverse
    if method == "right-to-left" or method == "bottom-to-top":
        reverse = True
  
    # handle if we are sorting against the y-coordinate rather than
    # the x-coordinate of the bounding box
    if method == "top-to-bottom" or method == "bottom-to-top":
        i = 1
  
    # construct the list of bounding boxes and sort them from top to
    # bottom
    boundingBoxes = [cv2.boundingRect(c) for c in cnts]
    (cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes),
        key=lambda b:b[1][i], reverse=reverse))
  
    # return the list of sorted contours and bounding boxes
    return (cnts, boundingBoxes)


def closest_node(pt, pt_list):
    
    min_dist_index = np.argmin(np.sum((np.array(pt_list) - np.array(pt))**2, axis=1))
    
    return min_dist_index
    
    

def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters, min_size):
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (width, height, n_channel) = image.shape
    
    #print("image shape: \n")
    #print(width, height, n_channel)
    
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    

    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    thresh_cleaned = clear_border(thresh)
    
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned_bw = clear_border(thresh)
    else:
        thresh_cleaned_bw = thresh
        
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)
    
    
    
    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    sizes = stats[1:, cv2.CC_STAT_AREA]
    
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    
    Coord_centroids = centroids
    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    nb_components = nb_components - 1
    
    
    
    #min_size = 70
    
    #max_size = width*height*0.1
    
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    #for every component in the image, keep it only if it's above min_size
    for i in range(0, nb_components):
        
        #print("{} nb_components found".format(i))
        '''
        if (sizes[i] >= min_size) and (Coord_left[i] > 1) and (Coord_top[i] > 1) and (Coord_width[i] - Coord_left[i] > 0) and (Coord_height[i] - Coord_top[i] > 0) and (centroids[i][0] - width*0.5 < 5) and ((centroids[i][1] - height*0.5 < 5)) and ((sizes[i] <= max_size)):
            img_thresh[output == i + 1] = 255
            
            print("Foreground center found ")
            
        elif ((Coord_width[i] - Coord_left[i])*0.5 - width < 5) and (centroids[i][0] - width*0.5 < 5) and (centroids[i][1] - height*0.5 < 5) and ((sizes[i] <= max_size)):
            imax = max(enumerate(sizes), key=(lambda x: x[1]))[0] + 1    
            #img_thresh[output == imax] = 255
            img_thresh[output == i + 1] = 255
            print("Foreground max found ")
        '''
        
        if (sizes[i] >= min_size) and (sizes[i] < max_size):
        
            img_thresh[output == i + 1] = 255
        
        
    
    #if mask contains mutiple non-conected parts, combine them into one. 
    contours, hier = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    #size_kernel = 13
    
    if len(contours) > 1:
        
        print("mask contains mutiple non-conected parts, combine them into one\n")
        
        kernel = np.ones((size_kernel,size_kernel), np.uint8)

        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        img_thresh = closing
    
    
    
    #from skimage import img_as_ubyte
    
    #img_thresh = img_as_ubyte(img_thresh)
    
    #print("img_thresh.dtype")
    #print(img_thresh.dtype)
    
    #return img_thresh
    return img_thresh
    
'''
def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)
    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis
'''





# Detect stickers in the image
def sticker_detect(img_ori, save_path):
    
    '''
    image_file_name = Path(image_file).name
    
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    print("Processing image : {0}\n".format(str(image_file)))
     
    # save folder construction
    mkpath = os.path.dirname(abs_path) +'/cropped'
    mkdir(mkpath)
    save_path = mkpath + '/'
    print ("results_folder: " + save_path)
    '''
   

    # load the image, clone it for output, and then convert it to grayscale
    #img_ori = cv2.imread(image_file)
    
    img_rgb = img_ori.copy()
      
    # Convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    
    #(minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(res)
    
    
    # Specify a threshold 
    threshold = 0.8
      
    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= threshold)  
    
    if len(loc):
    
        (y,x) = np.unravel_index(res.argmax(), res.shape)
    
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)
    
        print(y,x)
        
        print(min_val, max_val, min_loc, max_loc)
        
    
        (startX, startY) = max_loc
        endX = startX + template.shape[1]
        endY = startY + template.shape[0]
        
        '''
        # Draw a rectangle around the matched region. 
        for pt in zip(*loc[::-1]): 
            sticker_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,255), 2)
        '''
        
        sticker_crop_img = img_rgb[startY:endY, startX:endX]


    return  sticker_crop_img




def comp_external_contour(orig, thresh, save_path):
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
   
    img_height, img_width, img_channels = orig.shape
    
    #index = 1
    
    
    print("Number of contours: {}".format(len(contours)))
    
    '''
    list_of_pts = []
    
    if len(contours) > 1:
        
        
        for ctr in contours:
            
            list_of_pts += [pt[0] for pt in ctr]
    
        center_pt = np.array(list_of_pts).mean(axis = 0) # get origin
        
        clock_ang_dist = clockwise_angle_and_distance(center_pt) # set origin
        
        list_of_pts = sorted(list_of_pts, key=clock_ang_dist) # use to sort
        
        contours_joined = np.array(list_of_pts).reshape((-1,1,2)).astype(np.int32)
        
        
        
        kernel = np.ones((size_kernel,size_kernel), np.uint8)
        dilation = cv2.dilate(thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        trait_img = closing
    '''

    
    #grid initialization
    ####################################################################
    #number of rows
    nRows = args['nRows']
    #nRows = 6
    # Number of columns
    mCols = args['mCols']
    #mCols = 5
    
    #Dimensions of the image
    sizeX = img_width
    sizeY = img_height
    #print(img.shape)

    grid_center_label = []
    grid_center_coord = []
    
        # sort contours by area size in descending order
    #cntsSorted = sorted(contours, key=lambda x: cv2.contourArea(x), reverse = True)
    grid_center_label_rec = []
    
    for i in range(0, nRows):
        
        for j in range(0, mCols):
            
            #roi = orig[int(i*sizeY/nRows):int(i*sizeY/nRows) + int(sizeY/nRows), int(j*sizeX/mCols):int(j*sizeX/mCols) + int(sizeX/mCols)]
            
            x_center = int(j*sizeX/mCols) + int(sizeX/mCols/2)
            
            y_center = int(i*sizeY/nRows) + int(sizeY/nRows/2)
            
            #trait_img = cv2.putText(trait_img_bk, "#{}{}".format(i,j), (int(x_center), int(y_center)), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
            
            label = (i+1,j+1)
            grid_center_label.append(label)
            
            coord = (x_center, y_center)
            grid_center_coord.append(coord)
            
    
    ####################################################################
    trait_img_bk = orig.copy()
    
    trait_img = orig.copy()
    
    box_coord_rec = []
    
    merged_c_idx = []
    
    
    lab = cv2.cvtColor(orig.copy(), cv2.COLOR_BGR2LAB)
    
    i = 0
    
    cl = ColorLabeler()
    
    #for c in contours:
    
    for idx, c in enumerate(contours):
        
        #shape detection
        #################################################################
        '''
        # here we are ignoring first counter because 
        # findcontour function detects whole image as shape
        if i == 0:
            i = 1
            continue
        
        #hull = cv2.convexHull(c)
        
        # cv2.approxPloyDP() function to approximate the shape
        approx = cv2.approxPolyDP(c, 0.01 * cv2.arcLength(c, True), True)
        # using drawContours() function
        #trait_img = cv2.drawContours(orig, [c], 0, (0, 0, 255), 10)
  
        # finding center point of shape
        M = cv2.moments(c)
        
        if M['m00'] != 0.0:
            x = int(M['m10']/M['m00'])
            y = int(M['m01']/M['m00'])
  
        # putting shape name at center of each shape
        if len(approx) == 3:
            trait_img = cv2.putText(trait_img_bk, 'Triangle', (x, y), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
    
        elif len(approx) == 4:
            trait_img = cv2.putText(trait_img_bk, 'Quadrilateral', (x, y), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
    
        elif len(approx) == 5:
            trait_img = cv2.putText(trait_img_bk, 'Pentagon', (x, y), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
    
        elif len(approx) == 6:
            trait_img = cv2.putText(trait_img_bk, 'Hexagon', (x, y), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
    
        else:
            trait_img = cv2.putText(trait_img_bk, 'circle', (x, y), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
        '''
        ############################################################3
        '''
        # black image
        mask = np.zeros((img_height, img_width), dtype=np.uint8)
        
        # assign contours in white color
        mask_contour= cv2.drawContours(mask, [c], 0, 255, -1)
        
        masked_fg_contour = cv2.bitwise_and(orig, orig, mask = mask_contour)
        
        # Convert color space to LAB space and extract L channel
        L, A, B = cv2.split(cv2.cvtColor(masked_fg_contour.copy(), cv2.COLOR_BGR2LAB))
        '''
        
        #finding center point of shape
        M = cv2.moments(c)
            
        if M['m00'] != 0.0:
            x_c_center = int(M['m10']/M['m00'])
            y_c_center = int(M['m01']/M['m00'])

        #finding closest point among the grid points list ot the M coordinates
        idx_closest = closest_node((x_c_center,y_c_center), grid_center_coord)
        
        print("idx_closest = {}  {}".format(idx_closest, grid_center_label[idx_closest]))
        
        
        grid_label_str = ''.join([str(value) for value in grid_center_label[idx_closest]])
            
        if (grid_label_str in grid_center_label_rec):
        #if (grid_label_str in grid_center_label_rec) and cv2.contourArea(contours[grid_center_label_rec.index(grid_label_str)]) < cv2.contourArea(c):
                
            print("Repeat ROI {} idx = {} detected!".format(grid_label_str, idx))

            index_c = grid_center_label_rec.index(grid_label_str)

            mergred_c = np.vstack((contours[index_c], c))
            
            merged_c_idx.append(idx)
            
        else:
            
            mergred_c = c
                
        print("ROI {} detected ...\n".format(grid_label_str))
        
        
        
        color = cl.label(lab, mergred_c)

        
        #get the bounding rect
        (x, y, w, h) = cv2.boundingRect(mergred_c)
        
        ratio_bbx = min(w,h)/max(w,h)
        
        
        #save bounding box coordinates 
        '''
        rect = cv2.minAreaRect(c)
        box = cv2.boxPoints(rect)
        box = np.array(box, dtype="int")
        box_coord_flat = box.flatten()
        
        #print("bbox coordinates :{0}".format(box_coord_flat))
        
        
        box_coord = []
        for item in box_coord_flat:
            box_coord.append(item)
            
        box_coord_rec.append(box_coord)
        '''
        

        
        
        #if w>img_width*0.05 and h>img_height*0.05:
            
        if w>0 and h>0 and color == 'green':
            
            offset_w = int(w*0.25)
            offset_h = int(h*0.25)
            
            start_y = 0 if (y-offset_h) < 0 else (y-offset_h)           
            end_y = img_height if (y+h+offset_h) > img_height else (y+h+offset_h)
            start_x = 0 if (x-offset_w) < 0 else (x-offset_w)
            end_x = img_width if (x+w+offset_w > img_width) else (x+w+offset_w)
            

            # draw a green rectangle to visualize the bounding rect
            roi = orig[start_y : end_y, start_x : end_x]
            
            #roi = masked_fg_contour[start_y : end_y, start_x : end_x]
            
            
           
            
            #print("ROI {} detected ...".format(index))
            
            #result_file = (save_path +  str(format(index, "02")) + '.' + ext)
            
            result_file = (save_path +  str(format(grid_label_str)) + '.' + ext)
            
            cv2.imwrite(result_file, roi)
            
            #trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 3)
            
            #trait_img = cv2.drawContours(orig, c, -1, (0,255,255), -1)
            
            trait_img = cv2.rectangle(trait_img_bk, (x, y), (x+w, y+h), (255, 255, 0), 3)
            
            #trait_img = cv2.putText(trait_img_bk, "#{}".format(color), (int(x), int(y)), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
            trait_img = cv2.putText(trait_img_bk, "#{}".format(grid_label_str), (int(x), int(y)), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 0, 255), 10)
            
            trait_img = cv2.circle(trait_img_bk, (x, y), 3, (255, 0, 0), 3)
            
            trait_img = cv2.circle(trait_img_bk, (x+w, y+h), 3, (255, 0, 0), 3)
            
            #index+= 1
            
            grid_center_label_rec.append(grid_label_str)
            
            box_coord_rec.append((x,y, x,y+w, x+w,y+w, x+w,y))
        
            #print(x,y, x,y+w, x+w,y+w, x+w,y)
     
    #print(len(grid_center_label_rec))
    
    #print(len(box_coord_rec))
    
    box_coord_rec_merged = [j for i, j in enumerate(box_coord_rec) if i not in merged_c_idx]
    
    grid_center_label_rec_merged = [j for i, j in enumerate(grid_center_label_rec) if i not in merged_c_idx]
    
    grid_center_label_rec_merged = list(map(int, grid_center_label_rec_merged))
    
    #print((grid_center_label_rec_merged))
    
    #print(len(box_coord_rec_merged))
    
    sorted_grid_center_label_rec_merged = np.argsort(grid_center_label_rec_merged)
    
    #print((sorted_grid_center_label_rec_merged))
    
    #sort all lists according to sorted_grid_center_label_rec_merged order index
    box_coord_rec_merged[:] = [box_coord_rec_merged[i] for i in sorted_grid_center_label_rec_merged] 

    return trait_img, box_coord_rec_merged, grid_center_label_rec_merged





def segmentation(image_file):
    
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(image_file)
    
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    file_size = os.path.getsize(image_file)/MBFACTOR
    
    print("Segmenting image : {0} \n".format(str(filename)))
    
    #print("Base image : {0} \n".format(str(base_name)))
    
    # load original image
    image = cv2.imread(image_file)
    
    img_height, img_width, img_channels = image.shape
    
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    base_name = os.path.splitext(os.path.basename(filename))[0]
    # save folder construction
    mkpath = os.path.dirname(abs_path) +'/' + base_name
    mkdir(mkpath)
    save_path = mkpath + '/'
    
    
    mkpath_mask = os.path.dirname(abs_path) +'/' + base_name + '/mask'
    mkdir(mkpath_mask)
    save_path_mask = mkpath_mask + '/'
    
    print("results_folder: {0}\n".format(str(save_path)))  
    
    
    if (file_size > 5.0):
        print("It will take some time due to large file size {0} MB".format(str(int(file_size))))
    else:
        print("Segmenting plant object using automatic color clustering method... ")
    
    #make backup image
    orig = image.copy()
    
    
    # Convert color space to LAB space and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(orig.copy(), cv2.COLOR_BGR2LAB))
    
    # save Lab result
    result_file = (save_path_mask + base_name + '_L.' + ext)
    cv2.imwrite(result_file, L)
    
    # save Lab result
    result_file = (save_path_mask + base_name + '_A.' + ext)
    cv2.imwrite(result_file, A)
    
    # save Lab result
    result_file = (save_path_mask + base_name + '_B.' + ext)
    cv2.imwrite(result_file, B)
    
    
    #min_size = 2000

    #color clustering based plant object segmentation
    thresh = color_cluster_seg(orig, args_colorspace, args_channels, args_num_clusters, min_size)
    
    result_mask = save_path_mask + base_name + '_mask.' + ext
    
    cv2.imwrite(result_mask, thresh)
    
    
    #find external contour and segment image into small ROI based on each plant
    (trait_img, box_coord_rec, grid_center_label_rec) = comp_external_contour(image.copy(), thresh, save_path)
    
    #print("bbox coordinates :{0}".format(box_coord_rec))
    
    result_file = save_path_mask + base_name + '_label.' + ext
            
    cv2.imwrite(result_file, trait_img)
    
    
    #########################################################
    #validation purpose, can be removed 
    #write out box coordinates for validation
    #print("bbox coordinates :{0}".format((box_coord_rec)))
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    sheet_leaf = wb.create_sheet()
    '''
    sheet.cell(row = 1, column = 1).value = 'c1x'
    sheet.cell(row = 1, column = 2).value = 'c1y'
    sheet.cell(row = 1, column = 3).value = 'c2x'
    sheet.cell(row = 1, column = 4).value = 'c2y'
    sheet.cell(row = 1, column = 5).value = 'c3x'
    sheet.cell(row = 1, column = 6).value = 'c3y'
    sheet.cell(row = 1, column = 7).value = 'c4x'
    sheet.cell(row = 1, column = 8).value = 'c4y'
    '''
    for row in box_coord_rec:
        sheet.append(row)
   
    basename_repace = base_name.replace('_rgb','')
    #file name and path
    bbox_file = (args["path"] + basename_repace + '_bbox.xlsx')
    
    wb.save(bbox_file)
    
    bbox_file_csv = (args["path"] + basename_repace + '_bbox.csv')
    #convert xlsx to csv format
    wb = openpyxl.load_workbook(bbox_file)
    sh = wb.active # was .get_active_sheet()
    with open(bbox_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows: # generator; was sh.rows
            c.writerow([cell.value for cell in r])
            
        #################################################################end of validation file
    
    
    
    
    '''
    (sticker_crop_img) = sticker_detect(image.copy(), save_path)
    
    # save segmentation result
    #result_file = (save_path + base_name + 'sticker_matched.' + args['filetype'])
    #print(result_file)
    #cv2.imwrite(result_file, sticker_overlay)
    
    thresh_sticker = color_cluster_seg(sticker_crop_img.copy(), args_colorspace, args_channels, 4, min_size = 1000)
    trait_img_sticker = comp_external_contour(sticker_crop_img.copy(), thresh_sticker, save_path_sticker)
    result_file_sticker = save_path_sticker +  '_label.' + ext
    cv2.imwrite(result_file_sticker, trait_img_sticker)
    # save segmentation result
    result_file = (save_path_sticker + base_name + '_sticker_match.' + args['filetype'])
    #print(result_file)
    cv2.imwrite(result_file, sticker_crop_img)
    '''
    
    return thresh
    #trait_img
    
    
    

if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    #ap.add_argument('-i', '--image', required = True, help = 'Path to image file')
    ap.add_argument("-p", "--path", required = True,    help="path to image file")
    ap.add_argument("-ft", "--filetype", required=True,    help="Image filetype")
    
    ap.add_argument('-s', '--color-space', type =str, default ='lab', help='Color space to use: BGR (default), HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type = str, default='1', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type = int, default = 2,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-min', '--min_size', type = int, default = 100,  help = 'min size of object to be segmented.')
    ap.add_argument('-max', '--max_size', type = int, default = 10000000,  help = 'max size of object to be segmented.')
    ap.add_argument("-nr", "--nRows", required = False,  type = int,  default = 6, help="number of rows")
    ap.add_argument("-nc", "--mCols", required = False,  type = int,  default = 5, help="number of columns")
    args = vars(ap.parse_args())
    
    
    
    # setting path to model file
    file_path = args["path"]
    ext = args['filetype']
    
    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']
    min_size = args['min_size']
    max_size = args['max_size']

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    
    size_kernel = 3
    
    
   
    '''
    global  template
    template_path = "/home/suxing/smart_plant/marker_template/sticker_template.jpg"
    # Read the template 
    template = cv2.imread(template_path, 0) 
    print(template)
    '''
    #print((imgList))
    
    #current_img = imgList[0]
    
    #(thresh, trait_img) = segmentation(current_img)
    
    
    # get cpu number for parallel processing
    #agents = psutil.cpu_count()   
    agents = multiprocessing.cpu_count()
    print("Using {0} cores to perform parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result = pool.map(segmentation, imgList)
        pool.terminate()
    
    
    '''
    #loop execute
    for image in imgList:
        
        (thresh) = segmentation(image)
    '''
        
    #color clustering based plant object segmentation
    #thresh = color_cluster_seg(orig, args_colorspace, args_channels, args_num_clusters)
    
    # save segmentation result
    #result_file = (save_path + filename + '_seg' + file_extension)
    #print(filename)
    #cv2.imwrite(result_file, thresh)
    
    
    #find external contour 
    #trait_img = comp_external_contour(image.copy(),thresh, file_path)
    
    #save segmentation result
    #result_file = (save_path + filename + '_excontour' + file_extension)
    #cv2.imwrite(result_file, trait_img)
    
    
    #accquire medial axis of segmentation mask
    #image_medial_axis = medial_axis_image(thresh)

    # save medial axis result
    #result_file = (save_path + filename + '_medial_axis' + file_extension)
    #cv2.imwrite(result_file, img_as_ubyte(image_medial_axis))
    
    

    

    

 

