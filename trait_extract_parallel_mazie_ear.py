'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract plant shoot traits (larea, kernel_area_ratio, max_width, max_height, avg_curv, color_cluster) by paralell processing 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2022-09-29

USAGE:

time python3 trait_extract_parallel_mazie_ear.py -p ~/example/plant_test/seeds/image2/ -ft png -min 10 -md 10

'''

# import the necessary packages
import os
import glob
import utils

from collections import Counter

import argparse

from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.color import rgb2lab, deltaE_cie76
from skimage import morphology
from skimage.segmentation import clear_border, watershed
from skimage.measure import regionprops

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage
from scipy.interpolate import interp1d

from skan import skeleton_to_csgraph, Skeleton, summarize, draw

#from skan import sholl_analysis

#import networkx as nx

import imutils
from imutils import perspective

import numpy as np
import argparse
import cv2

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import math
import openpyxl
import csv
    
from tabulate import tabulate

import warnings
warnings.filterwarnings("ignore")

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from pathlib import Path 

from matplotlib import collections

from pylibdmtx.pylibdmtx import decode

import re


MBFACTOR = float(1<<20)

class ComputeCurvature:

    def __init__(self,x,y):
        """ Initialize some variables """
        self.xc = 0  # X-coordinate of circle center
        self.yc = 0  # Y-coordinate of circle center
        self.r = 0   # Radius of the circle
        self.xx = np.array([])  # Data points
        self.yy = np.array([])  # Data points
        self.x = x  # X-coordinate of circle center
        self.y = y  # Y-coordinate of circle center

    def calc_r(self, xc, yc):
        """ calculate the distance of each 2D points from the center (xc, yc) """
        return np.sqrt((self.xx-xc)**2 + (self.yy-yc)**2)

    def f(self, c):
        """ calculate the algebraic distance between the data points and the mean circle centered at c=(xc, yc) """
        ri = self.calc_r(*c)
        return ri - ri.mean()

    def df(self, c):
        """ Jacobian of f_2b
        The axis corresponding to derivatives must be coherent with the col_deriv option of leastsq"""
        xc, yc = c
        df_dc = np.empty((len(c), self.x.size))

        ri = self.calc_r(xc, yc)
        df_dc[0] = (xc - self.x)/ri                   # dR/dxc
        df_dc[1] = (yc - self.y)/ri                   # dR/dyc
        df_dc = df_dc - df_dc.mean(axis=1)[:, np.newaxis]
        return df_dc

    def fit(self, xx, yy):
        self.xx = xx
        self.yy = yy
        center_estimate = np.r_[np.mean(xx), np.mean(yy)]
        center = optimize.leastsq(self.f, center_estimate, Dfun=self.df, col_deriv=True)[0]

        self.xc, self.yc = center
        ri = self.calc_r(*center)
        self.r = ri.mean()

        return 1 / self.r  # Return the curvature


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False
        


def sort_contours(cnts, method="left-to-right"):
    
    # initialize the reverse flag and sort index
    reverse = False
    i = 0
    # handle if we need to sort in reverse
    if method == "right-to-left" or method == "bottom-to-top":
        reverse = True
    # handle if we are sorting against the y-coordinate rather than
    # the x-coordinate of the bounding box
    if method == "top-to-bottom" or method == "bottom-to-top":
        i = 1
    # construct the list of bounding boxes and sort them from top to
    # bottom
    boundingBoxes = [cv2.boundingRect(c) for c in cnts]
    (cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes), key=lambda b:b[1][i], reverse=reverse))
    # return the list of sorted contours and bounding boxes
    return cnts



def mutilple_objects_seg(orig):
    
    # load the image and perform pyramid mean shift filtering
    # to aid the thresholding step
    #Parse image path  and create result image path
    #path, filename = os.path.split(image_file)

    #print("processing image : {0} \n".format(str(filename)))

    #load the image and perform pyramid mean shift filtering to aid the thresholding step
    #image = cv2.imread(image_file)
    
    #orig = image.copy()
    
    #img_width, img_height, img_channels = orig.shape

    shifted = cv2.pyrMeanShiftFiltering(orig, 21, 70)

    height, width, channels = orig.shape

    # convert the mean shift image to grayscale, then apply
    # Otsu's thresholding
    gray = cv2.cvtColor(shifted, cv2.COLOR_BGR2GRAY)
    
    thresh = cv2.threshold(gray, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    # Taking a matrix of size 25 as the kernel
    kernel = np.ones((25,25), np.uint8)
    
    thresh_dilation = cv2.dilate(thresh, kernel, iterations=1)
    
    thresh_erosion = cv2.erode(thresh, kernel, iterations=1)
    
    #define result path for labeled images
    #result_img_path = save_path_label + str(filename[0:-4]) + '_thresh.jpg'
    #cv2.imwrite(result_img_path, thresh_erosion)
    
    # find contours in the thresholded image
    cnts = cv2.findContours(thresh_erosion.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    cnts = imutils.grab_contours(cnts)
    
    cnts_sorted = sorted(cnts, key=cv2.contourArea, reverse=True)[0:2]

    cnts_sorted = sort_contours(cnts_sorted, method="left-to-right")
    
    
    #c = max(cnts, key=cv2.contourArea)
    center_locX = []
    center_locY = []
    cnt_area = [0] * 2
    
    img_thresh = np.zeros(orig.shape, np.uint8)
    
    img_overlay_bk = orig
    
    
    
    for idx, c in enumerate(cnts_sorted):
        
        # compute the center of the contour
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        
        center_locX.append(cX)
        center_locY.append(cY)
        
        #cnt_area.append(cv2.contourArea(c))
        
        cnt_area[idx] = cv2.contourArea(c)
        
        # draw the contour and center of the shape on the image
        img_overlay = cv2.drawContours(img_overlay_bk, [c], -1, (0, 255, 0), 2)
        mask_seg = cv2.drawContours(img_thresh, [c], -1, (255,255,255),-1)
        #center_result = cv2.circle(img_thresh, (cX, cY), 14, (0, 0, 255), -1)
        img_overlay = cv2.putText(img_overlay_bk, "{}".format(idx +1), (cX - 20, cY - 20), cv2.FONT_HERSHEY_SIMPLEX, 5.5, (255, 0, 0), 5)
        
    
    #print(center_locX, center_locY)
    
    divide_X = int(sum(center_locX) / len(center_locX))
    divide_Y = int(sum(center_locY) / len(center_locY))
    
    #print(divide_X, divide_Y)
    
    
    #center_result = cv2.circle(image, (divide_X, divide_Y), 14, (0, 255, 0), -1)
    
    
    #define result path for labeled images
    #result_img_path = save_path_label + str(filename[0:-4]) + '_center.jpg'
    #cv2.imwrite(result_img_path, center_result)
    
    
    
    left_img = orig[0:height, 0:divide_X]
    #define result path for labeled images
    #result_img_path = save_path_label + str(filename[0:-4]) + '_left.jpg'
    #cv2.imwrite(result_img_path, left_img)
    
    
    right_img = orig[0:height, divide_X:width]
    #define result path for labeled images
    #result_img_path = save_path_label + str(filename[0:-4]) + '_right.jpg'
    #cv2.imwrite(result_img_path, right_img)
    
    mask_seg_gray = cv2.cvtColor(mask_seg, cv2.COLOR_BGR2GRAY)
    
    #(mask_seg_binary, im_bw) = cv2.threshold(mask_seg_gray, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    
    return left_img, right_img, mask_seg_gray, img_overlay, cnt_area



def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes
        
        
    #image = cv2.pyrMeanShiftFiltering(image, 21, 70)

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (width, height, n_channel) = image.shape
    
    #print("image shape: \n")
    #print(width, height, n_channel)
    
    
    
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    

    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    #thresh_cleaned = (thresh)
    
    
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned = clear_border(thresh)
    else:
        thresh_cleaned = thresh
    
     
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)

    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    sizes = stats[1:, cv2.CC_STAT_AREA]
    
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    
    Coord_centroids = centroids
    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    nb_components = nb_components - 1
    
    
    
    max_size = width*height*0.1
    
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    #for every component in the image, keep it only if it's above min_size
    for i in range(0, nb_components):
        
        '''
        #print("{} nb_components found".format(i))
        
        if (sizes[i] >= min_size) and (Coord_left[i] > 1) and (Coord_top[i] > 1) and (Coord_width[i] - Coord_left[i] > 0) and (Coord_height[i] - Coord_top[i] > 0) and (centroids[i][0] - width*0.5 < 10) and ((centroids[i][1] - height*0.5 < 10)) and ((sizes[i] <= max_size)):
            img_thresh[output == i + 1] = 255
            
            print("Foreground center found ")
            
        elif ((Coord_width[i] - Coord_left[i])*0.5 - width < 15) and (centroids[i][0] - width*0.5 < 15) and (centroids[i][1] - height*0.5 < 15) and ((sizes[i] <= max_size)):
            imax = max(enumerate(sizes), key=(lambda x: x[1]))[0] + 1    
            img_thresh[output == imax] = 255
            print("Foreground max found ")
        '''
        
        if (sizes[i] >= min_size):
        
            img_thresh[output == i + 1] = 255
        
    #from skimage import img_as_ubyte
    
    #img_thresh = img_as_ubyte(img_thresh)
    
    #print("img_thresh.dtype")
    #print(img_thresh.dtype)
    
    
    #if mask contains mutiple non-conected parts, combine them into one. 
    contours, hier = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) > 1:
        
        print("mask contains mutiple non-conected parts, combine them into one\n")
        
        kernel = np.ones((10,10), np.uint8)

        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        img_thresh = closing
        
    
    
    
    #return img_thresh
    return img_thresh
    
    #return thresh_cleaned
    

def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis


def skeleton_bw(thresh):

    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    #skeleton = morphology.skeletonize(image_bw)
    
    skeleton = morphology.thin(image_bw)

    skeleton_img = skeleton.astype(np.uint8) * 255

    return skeleton_img, skeleton

'''
def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels
'''


def percentage(part, whole):
  
  #percentage = "{:.0%}".format(float(part)/float(whole))
  
  percentage = "{:.2f}".format(float(part)/float(whole))
  
  return str(percentage)







def midpoint(ptA, ptB):
    return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)




def comp_external_contour(orig, thresh, img_overlay):
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
   
    #contours, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    
    img_height, img_width, img_channels = orig.shape
   
    index = 1
    
    trait_img = orig.copy()
    
    crop_img = orig.copy()
    
    area = 0
    kernel_area_ratio = 0
    w=h=0
    
    print("length of contours = {}\n".format(len(contours)))
    ####################################################################################
    
    
    contours = sorted(contours, key = cv2.contourArea, reverse = True)
    
    contours_sorted = sort_contours(contours, method="left-to-right")
    
    #c_max = max(contours, key = cv2.contourArea)
    
    area_c_cmax = 0
    
    area_holes_sum = 0
    
    cnt_area = [0] * 2
    
    cnt_width = []
    cnt_height = []
    
    orig = img_overlay
    ###########################################################################
    for index, c in enumerate(contours_sorted):
    #for c in contours:

        if index < 2:
            
            #get the bounding rect
            (x, y, w, h) = cv2.boundingRect(c)
            
            trait_img = cv2.drawContours(orig, c, -1, (255, 255, 0), 1)
            
            
            #roi = orig[y:y+h, x:x+w]
            
            print("ROI {} detected ...\n".format(index+1))
            
            # draw a green rectangle to visualize the bounding rect
            trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 4)
            
            '''
            ####################################################################
            margin = 150

            # define crop region
            start_y = int((y - margin) if (y - margin )> 0 else 0)

            start_x = int((x - margin) if (x - margin )> 0 else 0)

            crop_width = int((x + margin + w) if (x + margin + w) < img_width else (img_width))

            crop_height = int((y + margin + h) if (y + margin + h) < img_height else (img_height))

            crop_img = crop_img[start_y:crop_height, start_x:crop_width]

            ######################################################################
            '''
            
            # compute the center of the contour
            M = cv2.moments(c)
            
            cX = int(M["m10"] / M["m00"])
            cY = int(M["m01"] / M["m00"])
            
            # draw the center of the shape on the image
            #trait_img = cv2.circle(orig, (cX, cY), 7, (255, 255, 255), -1)
            #trait_img = cv2.putText(orig, "center", (cX - 20, cY - 20),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
            
            #################################################################################
            '''
            # compute the (rotated) bounding box of the contour to get the min area rect
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            
            # convert all coordinates floating point values to int
            box = np.int0(box)
            
            #draw a blue 'nghien' rectangle
            trait_img = cv2.drawContours(orig, [box], -1, (255, 0, 0), 3)
            
            # unpack the ordered bounding box, then compute the midpoint
            # between the top-left and top-right coordinates, followed by
            # the midpoint between bottom-left and bottom-right coordinates
            (tl, tr, br, bl) = box
            (tltrX, tltrY) = midpoint(tl, tr)
            (blbrX, blbrY) = midpoint(bl, br)
            
            # compute the midpoint between the top-left and top-right points,
            # followed by the midpoint between the top-righ and bottom-right
            (tlblX, tlblY) = midpoint(tl, bl)
            (trbrX, trbrY) = midpoint(tr, br)
            
            # draw the midpoints on the image
            trait_img = cv2.circle(orig, (int(tltrX), int(tltrY)), 5, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, (int(blbrX), int(blbrY)), 5, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, (int(tlblX), int(tlblY)), 5, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, (int(trbrX), int(trbrY)), 5, (255, 0, 0), -1)
            
            # draw lines between the midpoints
            trait_img = cv2.line(orig, (int(tltrX), int(tltrY)), (int(blbrX), int(blbrY)), (255, 0, 255), 2)
            trait_img = cv2.line(orig, (int(tlblX), int(tlblY)), (int(trbrX), int(trbrY)), (255, 0, 255), 2)
            
            
            # compute the Euclidean distance between the midpoints
            dA = dist.euclidean((tltrX, tltrY), (blbrX, blbrY))
            dB = dist.euclidean((tlblX, tlblY), (trbrX, trbrY))

            # draw the object sizes on the image
            #cv2.putText(orig, "{:.0f} pixels".format(dA), (int(tltrX), int(tltrY)), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
            #cv2.putText(orig, "{:.0f} pixels".format(dB), (int(trbrX), int(trbrY)), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
            '''
            ###############################################################################

            # the midpoint between bottom-left and bottom-right coordinates
            #(tl, tr, br, bl) = box
            
            tl = (x, y+h*0.5)
            tr = (x+w, y+h*0.5)
            br = (x+w*0.5, y)
            bl = (x+w*0.5, y+h)
            
            # compute the midpoint between bottom-left and bottom-right coordinates
            (tltrX, tltrY) = midpoint(tl, tr)
            (blbrX, blbrY) = midpoint(bl, br)
            
             # compute the midpoint between the top-left and top-right points,
            # followed by the midpoint between the top-righ and bottom-right
            #(tlblX, tlblY) = midpoint(tl, bl)
            #(trbrX, trbrY) = midpoint(tr, br)

            # draw the midpoints on the image
            trait_img = cv2.circle(orig, (int(tltrX), int(tltrY)), 15, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, (int(blbrX), int(blbrY)), 15, (255, 0, 0), -1)
            #trait_img = cv2.circle(orig, (int(tlblX), int(tlblY)), 5, (255, 0, 0), -1)
            #trait_img = cv2.circle(orig, (int(trbrX), int(trbrY)), 5, (255, 0, 0), -1)
            
            # draw lines between the midpoints
            trait_img = cv2.line(orig, (int(x), int(y+h*0.5)), (int(x+w), int(y+h*0.5)), (255, 0, 255), 6)
            trait_img = cv2.line(orig, (int(x+w*0.5), int(y)), (int(x+w*0.5), int(y+h)), (255, 0, 255), 6)
            

            hull = cv2.convexHull(c)
            
            # draw convexhull in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 2)
            
            area_c_cmax = cv2.contourArea(c)
            #hull_area = cv2.contourArea(hull)
            
            cnt_area[index] = (area_c_cmax)
            cnt_width.append(w)
            cnt_height.append(h)
            
            
            print("Contour {0} shape info: Width = {1:.2f}, height= {2:.2f}, area = {3:.2f}\n".format(index+1, w, h, area_c_cmax))
   
            
    return trait_img, cnt_area, cnt_width, cnt_height
    




def RGB2HEX(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))



'''
def color_quantization(image, mask, save_path, num_clusters):
    
    #grab image width and height
    (h, w) = image.shape[:2]
    
    #change the color storage order
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    #apply the mask to get the segmentation of plant
    masked_image = cv2.bitwise_and(image, image, mask = mask)
       
    # reshape the image to be a list of pixels
    pixels = masked_image.reshape((masked_image.shape[0] * masked_image.shape[1], 3))
        
    ############################################################
    #Clustering process
    ###############################################################
    # cluster the pixel intensities
    clt = MiniBatchKMeans(n_clusters = num_clusters)
    #clt = KMeans(n_clusters = args["clusters"])
    clt.fit(pixels)

    #assign labels to each cluster 
    labels = clt.fit_predict(pixels)

    #obtain the quantized clusters using each label
    quant = clt.cluster_centers_.astype("uint8")[labels]

    # reshape the feature vectors to images
    quant = quant.reshape((h, w, 3))
    image_rec = pixels.reshape((h, w, 3))
    
    # convert from L*a*b* to RGB
    quant = cv2.cvtColor(quant, cv2.COLOR_RGB2BGR)
    image_rec = cv2.cvtColor(image_rec, cv2.COLOR_RGB2BGR)
    
    # display the images 
    #cv2.imshow("image", np.hstack([image_rec, quant]))
    #cv2.waitKey(0)
    
    #define result path for labeled images
    result_img_path = save_path + 'cluster_out.png'
    
    # save color_quantization results
    cv2.imwrite(result_img_path,quant)

    #Get colors and analze them from masked image
    counts = Counter(labels)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = clt.cluster_centers_
    
    #print(type(center_colors))

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]
    
    
    #######################################################################################
    threshold = 60
    
    selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))

    for i in range(num_clusters):
        curr_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[i]]]))) 
        diff = deltaE_cie76(selected_color, curr_color)
        if (diff < threshold):
            print("Color difference value is : {0} \n".format(str(diff)))
    ###########################################################################################
    #print(hex_colors)
    
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    
    #print(index_bkg[0])

    #print(counts)
    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)
        
    # build a histogram of clusters and then create a figure representing the number of pixels labeled to each color
    hist = utils.centroid_histogram(clt)

    # remove the background color cluster
    clt.cluster_centers_ = np.delete(clt.cluster_centers_, index_bkg[0], axis=0)
    
    #build a histogram of clusters using center lables
    numLabels = utils.plot_centroid_histogram(save_path,clt)

    #create a figure representing the distribution of each color
    bar = utils.plot_colors(hist, clt.cluster_centers_)

    #save a figure of color bar 
    utils.plot_color_bar(save_path, bar)

    return rgb_colors
'''

def get_cmap(n, name = 'hsv'):
    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    return plt.cm.get_cmap(name, n)
    
    

def color_region(image, mask, save_path, num_clusters):
    
    # read the image
     #grab image width and height
    (h, w) = image.shape[:2]

    #apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask = mask)
    
    #define result path for labeled images
    result_img_path = save_path + 'masked.png'
    cv2.imwrite(result_img_path, masked_image_ori)
    
    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))
    
    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    #num_clusters = 5
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)


    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    #define result path for labeled images
    result_img_path = save_path + 'clustered.png'
    cv2.imwrite(result_img_path, segmented_image_BRG)


    '''
    fig = plt.figure()
    ax = Axes3D(fig)        
    for label, pix in zip(labels, segmented_image):
        ax.scatter(pix[0], pix[1], pix[2], color = (centers))
            
    result_file = (save_path + base_name + 'color_cluster_distributation.png')
    plt.savefig(result_file)
    '''
    #Show only one chosen cluster 
    #masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    #cluster = 2

    cmap = get_cmap(num_clusters+1)
    
    #clrs = sns.color_palette('husl', n_colors = num_clusters)  # a list of RGB tuples

    color_conversion = interp1d([0,1],[0,255])


    for cluster in range(num_clusters):

        print("Processing color cluster {0} ...\n".format(cluster))
        #print(clrs[cluster])
        #print(color_conversion(clrs[cluster]))

        masked_image[labels_flat == cluster] = centers[cluster]

        #print(centers[cluster])

        #convert back to original shape
        masked_image_rp = masked_image.reshape(image_RGB.shape)

        #masked_image_BRG = cv2.cvtColor(masked_image, cv2.COLOR_RGB2BGR)
        #cv2.imwrite('maksed.png', masked_image_BRG)

        gray = cv2.cvtColor(masked_image_rp, cv2.COLOR_BGR2GRAY)

        # threshold the image, then perform a series of erosions +
        # dilations to remove any small regions of noise
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)[1]

        #thresh_cleaned = clear_border(thresh)

        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        #c = max(cnts, key=cv2.contourArea)

        '''
        # compute the center of the contour area and draw a circle representing the center
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        # draw the countour number on the image
        result = cv2.putText(masked_image_rp, "#{}".format(cluster + 1), (cX - 20, cY), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
        '''
        
        
        if not cnts:
            print("findContours is empty")
        else:
            
            # loop over the (unsorted) contours and draw them
            for (i, c) in enumerate(cnts):

                result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(np.random.random(3)), 2)
                #result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(clrs[cluster]), 2)

            #result = result(np.where(result == 0)== 255)
            result[result == 0] = 255


            result_BRG = cv2.cvtColor(result, cv2.COLOR_RGB2BGR)
            result_img_path = save_path + 'result_' + str(cluster) + '.png'
            cv2.imwrite(result_img_path, result_BRG)


    
    counts = Counter(labels_flat)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = centers

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]

    #print(hex_colors)
    
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    
    #print(index_bkg[0])

    #print(counts)
    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)

   
    return rgb_colors, counts, hex_colors


def _normalise_image(image, *, image_cmap=None):
    image = img_as_float(image)
    if image.ndim == 2:
        if image_cmap is None:
            image = gray2rgb(image)
        else:
            image = plt.get_cmap(image_cmap)(image)[..., :3]
    return image







'''
def overlay_skeleton_endpoints(image, stats, *, image_cmap=None, axes=None):

    image = _normalise_image(image, image_cmap=image_cmap)
    summary = stats
    # transforming from row, col to x, y
    #coords_cols = (['image-coord-src-%i' % i for i in [1, 0]] +
    #               ['image-coord-dst-%i' % i for i in [1, 0]])
    
    coords_cols_src = (['image-coord-src-%i' % i for i in [1, 0]])
    coords_cols_dst = (['image-coord-dst-%i' % i for i in [1, 0]])
    
    #coords = summary[coords_cols].values.reshape((-1, 1, 2))
    
    coords_src = summary[coords_cols_src].values
    coords_dst = summary[coords_cols_dst].values

    coords_src_x = [i[0] for i in coords_src]
    coords_src_y = [i[1] for i in coords_src]
    
    coords_dst_x = [i[0] for i in coords_dst]
    coords_dst_y = [i[1] for i in coords_dst]
    
    img_marker = np.zeros_like(image, dtype = np.uint8)
    img_marker.fill(0) # or img[:] = 255
    img_marker[list(map(int, coords_src_y)), list(map(int, coords_src_x))] = 255
    img_marker[list(map(int, coords_dst_y)), list(map(int, coords_dst_x))] = 255
    
    #print("img_marker")
    #print(img_marker.shape)
    
    if axes is None:
        fig, axes = plt.subplots()
    
    axes.axis('off')
    axes.imshow(image)

    axes.scatter(coords_src_x, coords_src_y, c = 'w')
    axes.scatter(coords_dst_x, coords_dst_y, c = 'w')

    return fig, img_marker
    #return coords
'''

def outlier_doubleMAD(data,thresh = 3.5):
    
    """
    FOR ASSYMMETRIC DISTRIBUTION
    Returns : filtered array excluding the outliers

    Parameters : the actual data Points array

    Calculates median to divide data into 2 halves.(skew conditions handled)
    Then those two halves are treated as separate data with calculation same as for symmetric distribution.(first answer) 
    Only difference being , the thresholds are now the median distance of the right and left median with the actual data median
    """
    
    # warning: this function does not check for NAs
    # nor does it address issues when 
    # more than 50% of your data have identical values
    m = np.median(data)
    abs_dev = np.abs(data - m)
    left_mad = np.median(abs_dev[data <= m])
    right_mad = np.median(abs_dev[data >= m])
    data_mad = left_mad * np.ones(len(data))
    data_mad[data > m] = right_mad
    modified_z_score = 0.6745 * abs_dev / data_mad
    modified_z_score[data == m] = 0
    return modified_z_score > thresh



# Read barcode in the imageand decode barcode info
def barcode_detect(img_ori):
    
    height, width = img_ori.shape[:2]
    
    barcode_info = decode((img_ori.tobytes(), width, height))
    
    if len(barcode_info) > 0:
    
        barcode_str = " ".join(str(x) for x in barcode_info)
        
        #print((barcode_str))
            
        tag_info = re.findall(r"'(.*?)'", barcode_str, re.DOTALL)
        
        tag_info = " ".join(str(x) for x in tag_info)
        
        tag_info = tag_info.replace("'", "")
        
        print("Tag info: {}\n".format(tag_info))
    
    else:
        
        print("barcode inforamation was not readable!\n")
        tag_info = 'Unreadable'
        
    return tag_info
    
    


# Detect marker in the image
def marker_detect(img_ori, template, selection_threshold):
    
    # load the image, clone it for output
    img_rgb = img_ori.copy()
      
    # Convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    
    #(minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(res)
    
    
    # Specify a threshold for template detection, can be adjusted
    #threshold = 0.8
      
    # Store the coordinates of matched area in a numpy array 
    #loc = np.where( res >= threshold)
    loc = np.where( res >= selection_threshold)   
    
    if len(loc):
    
        (y,x) = np.unravel_index(res.argmax(), res.shape)
    
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)

        (startX, startY) = max_loc
        endX = startX + template.shape[1]
        endY = startY + template.shape[0]
        
        
        marker_img = img_ori[startY:endY, startX:endX]
        
        marker_overlay = marker_img
        
        
        
        # Draw a rectangle around the matched region. 
        #for pt in zip(*loc[::-1]): 
            
            #marker_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,0), 1)

    
        # load the marker image, convert it to grayscale
        marker_img_gray = cv2.cvtColor(marker_img, cv2.COLOR_BGR2GRAY) 
    
       
        # load the image and perform pyramid mean shift filtering to aid the thresholding step
        shifted = cv2.pyrMeanShiftFiltering(marker_img, 21, 51)
        
        # convert the mean shift image to grayscale, then apply Otsu's thresholding
        gray = cv2.cvtColor(shifted, cv2.COLOR_BGR2GRAY)
        
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        
        # detect contours in the mask and grab the largest one
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        cnts = imutils.grab_contours(cnts)
        
        largest_cnt = max(cnts, key=cv2.contourArea)
        
        print("[INFO] {} unique contours found in marker_img".format(len(cnts)))
        
        # compute the radius of the detected coin
        # calculate the center of the contour
        M = cv2.moments(largest_cnt )
        x = int(M["m10"] / M["m00"])
        y = int(M["m01"] / M["m00"])

        # calculate the radius of the contour from area (I suppose it's a circle)
        area = cv2.contourArea(largest_cnt)
        radius = np.sqrt(area/math.pi)
        coins_width_contour = 2* radius
    
        print("The width of Brazil 1 Real coin in the marker image is {:.0f} pixels\n".format(coins_width_contour))
        
        
        # draw a circle enclosing the object
        ((x, y), radius) = cv2.minEnclosingCircle(largest_cnt) 
        coins_width_circle = 2* radius
        
        # Brazil 1 Real coin dimension is 27 × 27 mm
        print("The width of Brazil 1 Real coin in the marker image is {:.0f} pixels\n".format(coins_width_circle))
        

        

    return  marker_img, thresh, coins_width_contour, coins_width_circle




# Convert it to LAB color space to access the luminous channel which is independent of colors.
def isbright(image_file):
    
    # Set up threshold value for luminous channel, can be adjusted and generalized 
    thresh = 1.5
    
    # Load image file 
    orig = cv2.imread(image_file)
    
    # Make backup image
    image = orig.copy()
    
    # Get file name
    #abs_path = os.path.abspath(image_file)
    
    #filename, file_extension = os.path.splitext(abs_path)
    #base_name = os.path.splitext(os.path.basename(filename))[0]

    image_file_name = Path(image_file).name
    
    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    
    # Normalize L channel by dividing all pixel values with maximum pixel value
    L = L/np.max(L)
    
    text_bool = "bright" if np.mean(L) < thresh else "dark"
    
    #return image_file_name, np.mean(L), text_bool
    
    print("np.mean(L) < thresh = {}".format(np.mean(L)))
    
    return np.mean(L) < thresh






def extract_traits(image_file):


    #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    base_name = os.path.splitext(os.path.basename(filename))[0]

    file_size = os.path.getsize(image_file)/MBFACTOR
    
    image_file_name = Path(image_file).name
   
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    
    print("Exacting traits for image : {0}\n".format(str(image_file_name)))
     
    # save folder construction
    if (args['result']):
        save_path = args['result']
    else:
        mkpath = os.path.dirname(abs_path) +'/' + base_name
        mkdir(mkpath)
        save_path = mkpath + '/'
        
        #track_save_path = os.path.dirname(abs_path) + '/trace/'
        #mkdir(track_save_path)

    print ("results_folder: " + save_path)
    
    #print ("track_save_path: " + track_save_path)
    
    
    
    
   
    
        
    if isbright(image_file):
    
        if (file_size > 5.0):
            print("It will take some time due to larger file size {0} MB".format(str(int(file_size))))
        else:
            print("Segmentaing plant object using automatic color clustering method... ")
        
        image = cv2.imread(image_file)
        
        #make backup image
        orig = image.copy()
        
        source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
        
        
        (left_img, right_img, mask_seg, img_overlay, cnt_area_internal) = mutilple_objects_seg(orig)
        # save segmentation result
        #result_file = (save_path + base_name + '_left' + file_extension)
        #cv2.imwrite(result_file, left_img)
        
        # save segmentation result
        #result_file = (save_path + base_name + '_right' + file_extension)
        #cv2.imwrite(result_file, right_img)
        
        # save segmentation result
        #result_file = (save_path + base_name + '_mask_seg' + file_extension)
        #cv2.imwrite(result_file, mask_seg)
        
        # save segmentation result
        result_file = (save_path + base_name + '_overlay' + file_extension)
        cv2.imwrite(result_file, img_overlay)
        
         
        args_colorspace = args['color_space']
        args_channels = args['channels']
        args_num_clusters = args['num_clusters']
        
        #color clustering based plant object segmentation
        thresh = color_cluster_seg(image.copy(), args_colorspace, args_channels, args_num_clusters)
        # save segmentation result
        #result_file = (save_path + base_name + '_color_cluster_seg' + file_extension)
        #cv2.imwrite(result_file, thresh)
        
        
       ###############################################################################################
        #combine two masks
        combined_mask = mask_seg | thresh
        
        
        # Taking a matrix of size 25 as the kernel
        kernel = np.ones((25,25), np.uint8)
        combined_mask = cv2.dilate(combined_mask, kernel, iterations=1)
        
        result_file = (save_path + base_name + '_combined_mask' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, combined_mask)
        
        
        thresh_combined_mask = cv2.threshold(combined_mask, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        # find contours in the thresholded image
        cnts = cv2.findContours(thresh_combined_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)

        cnts_sorted = sorted(cnts, key=cv2.contourArea, reverse=True)
        
        for c in cnts_sorted[0:2]:

            # draw the contour and center of the shape on the image
            img_overlay = cv2.drawContours(img_overlay, [c], -1, (0, 255, 0), 2)
        
        # save segmentation result
        result_file = (save_path + base_name + '_overlay_combined' + file_extension)
        cv2.imwrite(result_file, img_overlay)
        
        ################################################################################################################################
        #find external contour 
        (trait_img, cnt_area_external, cnt_width, cnt_height) = comp_external_contour(image.copy(), thresh_combined_mask, img_overlay)
        
        # save segmentation result
        result_file = (save_path + base_name + '_excontour' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, trait_img)
        
        
        #print(cnt_area_internal)
        #print(cnt_area_external)
        

        #compute the area ratio of interal contour verse external contour, kernal area ratio
        area_max_external = max(cnt_area_external)
        
        sum_area_internal = sum(cnt_area_internal)
        

        if (cnt_area_internal[0] >= cnt_area_external[0])  or  (cnt_area_internal[1] >= cnt_area_external[1]):
            
            if sum_area_internal < area_max_external:
                
                cnt_area_external[0] = cnt_area_external[1] = np.mean(cnt_area_external)
            else:
                cnt_area_external[0] = cnt_area_external[1] = area_max_external

        
        area_ratio = (cnt_area_internal[0]/cnt_area_external[0], cnt_area_internal[1]/cnt_area_external[1])
        
        
        area_ratio  = [ elem for elem in area_ratio if elem < 1]
        
        '''
        print("[INFO]: The two kernel_area = {}".format(area_ratio))
        print("[INFO]: The two width = {}".format(cnt_width))
        print("[INFO]: The two height = {}".format(cnt_height))
        '''
        kernel_area_ratio = np.mean(area_ratio)
        
        kernel_area = sum(cnt_area_internal) / len(cnt_area_internal)
        
        max_width = sum(cnt_width) / len(cnt_width)
        
        max_height = sum(cnt_height) / len(cnt_height)
        
        
        ################################################################################################
        '''
        num_clusters = 5
        
        #save color analysisi/quantization result
        (rgb_colors, counts, hex_colors) = color_region(left_img.copy(), thresh, save_path, num_clusters)
        

        #print("hex_colors = {} {}\n".format(hex_colors, type(hex_colors)))
        
        list_counts = list(counts.values())
        
        #list_hex_colors = list(hex_colors)
        
        #print(type(list_counts))
        
        color_ratio = []
        
        for value_counts, value_hex in zip(list_counts, hex_colors):
            
            #print(percentage(value, np.sum(list_counts)))
            
            color_ratio.append(percentage(value_counts, np.sum(list_counts)))
            
            #print("value_hex = {0}".format(value_hex))
            
            #value_hex.append(value_hex)
            
            
        #color_ratio_rec.append(color_ratio)
        #color_value_rec.append(hex_colors)
        
        
        
        selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))
        '''
        
        ####################################################
        '''
        print("Color difference are : ") 
        
        print(selected_color)
        
        color_diff = []
        
        for index, value in enumerate(rgb_colors): 
            #print(index, value) 
            curr_color = rgb2lab(np.uint8(np.asarray([[value]])))
            diff = deltaE_cie76(selected_color, curr_color)
            
            color_diff.append(diff)
            
            print(index, value, diff) 
        
        '''
        '''
        ###############################################
        
        #accquire medial axis of segmentation mask
        #image_skeleton = medial_axis_image(thresh)
        
        image_skeleton, skeleton = skeleton_bw(thresh)

        # save _skeleton result
        result_file = (save_path + base_name + '_skeleton' + file_extension)
        cv2.imwrite(result_file, img_as_ubyte(image_skeleton))
        
        
        ###
        # ['skeleton-id', 'node-id-src', 'node-id-dst', 'branch-distance', 
        #'branch-type', 'mean-pixel-value', 'stdev-pixel-value', 
        #'image-coord-src-0', 'image-coord-src-1', 'image-coord-dst-0', 'image-coord-dst-1', 
        #'coord-src-0', 'coord-src-1', 'coord-dst-0', 'coord-dst-1', 'euclidean-distance']
        ###
        
        #get brach data
        branch_data = summarize(Skeleton(image_skeleton))
        #print(branch_data)
        
        #select end branch
        sub_branch = branch_data.loc[branch_data['branch-type'] == 1]
        
        sub_branch_branch_distance = sub_branch["branch-distance"].tolist()
     
        # remove outliers in branch distance 
        outlier_list = outlier_doubleMAD(sub_branch_branch_distance, thresh = 3.5)
        
        indices = [i for i, x in enumerate(outlier_list) if x]
        
        sub_branch_cleaned = sub_branch.drop(sub_branch.index[indices])

        #print(outlier_list)
        
        #print(indices)
        
        #print(sub_branch)
        
        #print(sub_branch_cleaned)
        
        #print(sub_branch_cleaned.iloc[:, 0])
        
        
        end_points_coord_y = sub_branch_cleaned["image-coord-src-0"].tolist()
        end_points_coord_x = sub_branch_cleaned["image-coord-src-1"].tolist()
        
        #skeleton_id_list = sub_branch_cleaned["skeleton-id"].tolist()
        
        maxpos_y = end_points_coord_y.index(max(end_points_coord_y))
        
        #print("[INFO] There are {} branch end points were found\n".format(end_points_coord_x))
        #print("[INFO] There are {} branch end points were found\n".format(end_points_coord_y))
        
        
        for index, (x,y) in enumerate(zip(end_points_coord_x, end_points_coord_y)):
            
            if index == maxpos_y:
                
                # draw the endpoints on the image
                end_point_overlay = cv2.circle(orig, (int(x), int(y)), 3, (0, 0, 255), -1)
                end_point_overlay = cv2.putText(orig, "{}".format("root point"), (int(x-25), int(y+15)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 0, 255), 2)
            
            else:
            
                # draw the endpoints on the image
                end_point_overlay = cv2.circle(orig, (int(x), int(y)), 2, (255, 0, 0), -1)
                end_point_overlay = cv2.putText(orig, "{}".format(index), (int(x), int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 0, 255), 2)
        
        
        result_file = (save_path + base_name + '_end_point_overlay' + file_extension)
        cv2.imwrite(result_file, end_point_overlay)
        '''

        '''
        # define the center/soma
        center = np.array([end_points_coord_x[maxpos_y], end_points_coord_y[maxpos_y]])
        
        # define radii at which to measure crossings
        radii = np.arange(4, 45, 4)
        
        # perform sholl analysis
        center, radii, counts = sholl_analysis(Skeleton(image_skeleton), center=center, shells=radii)
        
        table = pd.DataFrame({'radius': radii, 'crossings': counts})
        
        print(table)
        '''

        '''
        branch_type_list = sub_branch_cleaned["branch-type"].tolist()
        
        #print(branch_type_list.count(1))
        
        print("[INFO] {} branch end points found\n".format(branch_type_list.count(1)))
        
        #img_hist = branch_data.hist(column = 'branch-distance', by = 'branch-type', bins = 100)
        #result_file = (save_path + base_name + '_hist' + file_extension)
        #plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        #plt.close()

        
        fig = plt.plot()
        
        source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
        
        img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-distance', skeleton_colormap = 'hsv')
        
        #img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-type', skeleton_colormap = 'hsv')
        
        result_file = (save_path + base_name + '_euclidean_graph_overlay' + file_extension)
        
        plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        
        plt.close()
        '''
        
        # detect the coin based on the template image
        (marker_coin_img, thresh_coin, coins_width_contour, coins_width_circle) = marker_detect(image.copy(), tp_coin, 0.8)
        
        # save segmentation result
        result_file = (save_path + base_name + '_coin_deteced.' + file_extension)
        cv2.imwrite(result_file, marker_coin_img)
        
        
        # detect the barcode based on the template image
        (marker_barcode_img, thresh_barcode, barcode_width_contour, barcode_width_circle) = marker_detect(image.copy(), tp_barcode, 0.6)

        # save segmentation result
        result_file = (save_path + base_name + '_barcode_deteced.' + file_extension)
        cv2.imwrite(result_file, marker_barcode_img)
        
        tag_info = barcode_detect(marker_barcode_img)
        

        ############################################## leaf number computation
        '''
        #min_distance_value = 3
            
        print("min_distance_value = {}\n".format(min_distance_value))
        
        #watershed based leaf area segmentaiton 
        labels = watershed_seg(orig, thresh, min_distance_value)
        
        #n_leaves = int(len(np.unique(labels)))
        

        
        #labels = watershed_seg_marker(orig, thresh, min_distance_value, img_marker)
        
        #individual_object_seg(orig, labels, save_path, base_name, file_extension)

        #save watershed result label image
        #Map component labels to hue val
        label_hue = np.uint8(128*labels/np.max(labels))
        #label_hue[labels == largest_label] = np.uint8(15)
        blank_ch = 255*np.ones_like(label_hue)
        labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

        # cvt to BGR for display
        labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

        # set background label to black
        labeled_img[label_hue==0] = 0
        result_file = (save_path + base_name + '_label' + file_extension)
        #plt.imsave(result_file, img_as_float(labels), cmap = "Spectral")
        cv2.imwrite(result_file, labeled_img)
        '''
        
        
        '''
        (avg_curv, label_trait, track_trait, leaf_index_rec, contours_rec, area_rec, curv_rec, kernel_area_ratio_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec, box_coord_rec) = leaf_traits_computation(orig, labels, save_path, base_name, file_extension)
        '''
        
        '''
        #########################################################validation purpose, can be removed 
        #write out box coordinates for validation
        #print("bbox coordinates :{0}".format((box_coord_rec)))
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'c1x'
        sheet.cell(row = 1, column = 2).value = 'c1y'
        sheet.cell(row = 1, column = 3).value = 'c2x'
        sheet.cell(row = 1, column = 4).value = 'c2y'
        sheet.cell(row = 1, column = 5).value = 'c3x'
        sheet.cell(row = 1, column = 6).value = 'c3y'
        sheet.cell(row = 1, column = 7).value = 'c4x'
        sheet.cell(row = 1, column = 8).value = 'c4y'
        
        for row in box_coord_rec:
            sheet.append(row)
       
        #file name and path
        bbox_file = (args["path"] + 'bbox.xlsx')
        
        wb.save(bbox_file)
        
        bbox_file_csv = (args["path"] + 'bbox.csv')
        #convert xlsx to csv format
        wb = openpyxl.load_workbook(bbox_file)
        sh = wb.active # was .get_active_sheet()
        with open(bbox_file_csv, 'w', newline = "") as f:
            c = csv.writer(f)
            for r in sh.rows: # generator; was sh.rows
                c.writerow([cell.value for cell in r])
        '''
        #################################################################end of validation file
        
        #n_leaves = int(len((leaf_index_rec)))
        
        #print('number of leaves{0}'.format(n_leaves))
        
        #save watershed result label image
        #result_file = (save_path + base_name + '_leafspec' + file_extension)
        #cv2.imwrite(result_file, label_trait)
        
        #save watershed result label image
        #result_file = (track_save_path + base_name + '_trace' + file_extension)
        #cv2.imwrite(result_file, track_trait)
        

        
        
    
    else:
        
        area = kernel_area_ratio = max_width = max_height = avg_curv = n_leaves = 0
        
    #print("[INFO] {} n_leaves found\n".format(len(np.unique(labels)) - 1))
    
    #Path("/tmp/d/a.dat").name
    
    #print("color_ratio = {}".format(color_ratio))
    
    #print("hex_colors = {}".format(hex_colors))
    
    #return image_file_name, area, kernel_area_ratio, max_width, max_height, avg_curv, n_leaves, color_ratio, hex_colors, leaf_index_rec, area_rec, curv_rec, kernel_area_ratio_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec
    
    #return image_file_name, area, kernel_area_ratio, max_width, max_height, color_ratio, hex_colors
    
    return image_file_name, tag_info, kernel_area, kernel_area_ratio, max_width, max_height
    





if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required = True,    help = "Image filetype")
    ap.add_argument('-mk', '--marker', required = False,  default ='/marker_template/marker.png',  help = "Marker file name")
    ap.add_argument('-bc', '--barcode', required = False,  default ='/marker_template/barcode.png',  help = "Barcode file name")
    ap.add_argument("-r", "--result", required = False,    help="result path")
    ap.add_argument('-s', '--color-space', type = str, required = False, default ='lab', help='Color space to use: BGR (default), HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type = str, required = False, default='1', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type = int, required = False, default = 2,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-min', '--min_size', type = int, required = False, default = 100,  help = 'min size of object to be segmented.')
    ap.add_argument('-md', '--min_dist', type = int, required = False, default = 10,  help = 'distance threshold of watershed segmentation.')
    
    
    args = vars(ap.parse_args())
    
    
    # setting path to model file
    file_path = args["path"]
    ext = args['filetype']
    
    coin_path = args["marker"]
    barcode_path = args["barcode"]
    
    min_size = args['min_size']
    min_distance_value = args['min_dist']
    
    
    
    global  tp_coin, tp_barcode
    # path of the marker (coin), default path will be '/marker_template/marker.png' and '/marker_template/barcode.png'
    # can be changed based on requirement
    
    template_path = file_path + coin_path
    barcode_path = file_path + barcode_path
    
    try:
        # check to see if file is readable
        with open(template_path) as tempFile:

            # Read the template 
            tp_coin = cv2.imread(template_path, 0)
            print("Template loaded successfully")
            
    except IOError as err:
        
        print("Error reading the Template file {0}: {1}".format(template_path, err))
        
        exit(0)

    
    try:
        # check to see if file is readable
        with open(barcode_path) as tempFile:

            # Read the template 
            tp_barcode = cv2.imread(barcode_path, 0)
            print("Barcode loaded successfully")
            
    except IOError as err:
        
        print("Error reading the Barcode file {0}: {1}".format(barcode_path, err))
        
        exit(0)
    
    

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))

    #print((imgList))
    #global save_path
    
    n_images = len(imgList)
    
    result_list = []
    
    #result_list_leaf = []
    
    
    #loop execute
    for image in imgList:
        
        #(filename, area, kernel_area_ratio, max_width, max_height, color_ratio, hex_colors) = extract_traits(image)
        
        (filename, tag_info, kernel_area, kernel_area_ratio, max_width, max_height) = extract_traits(image)
        
        #result_list.append([filename, area, kernel_area_ratio, max_width, max_height, color_ratio[0], color_ratio[1], color_ratio[2], color_ratio[3], hex_colors[0], hex_colors[1], hex_colors[2], hex_colors[3]])
        
        result_list.append([filename, tag_info, kernel_area, kernel_area_ratio, max_width, max_height])

        #print(leaf_color_value_rec)
        
        #for i in range(len(leaf_index_rec)):
            
            #result_list_leaf.append([filename, leaf_index_rec[i], area_rec[i], curv_rec[i], kernel_area_ratio_rec[i], major_axis_rec[i], minor_axis_rec[i], leaf_color_ratio_rec[i][0], leaf_color_ratio_rec[i][1], leaf_color_ratio_rec[i][2], leaf_color_ratio_rec[i][3], leaf_color_value_rec[i][0],leaf_color_value_rec[i][1],leaf_color_value_rec[i][2],leaf_color_value_rec[i][3]])
    '''
    
    #print(result_list)
    
    for image in imgList:
    
        extract_traits(image)
    '''

    '''
    # get cpu number for parallel processing
    agents = psutil.cpu_count()   
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result_list = pool.map(extract_traits, imgList)
        pool.terminate()
    '''
    
    
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
 
    #table = tabulate(result_list, headers = ['filename', 'mazie_ear_area', 'kernel_area_ratio', 'max_width', 'max_height' ,'cluster 1', 'cluster 2', 'cluster 3', 'cluster 4', 'cluster 1 hex value', 'cluster 2 hex value', 'cluster 3 hex value', 'cluster 4 hex value'], tablefmt = 'orgtbl')
    table = tabulate(result_list, headers = ['filename', 'tag_info', 'avg_kernel_area', 'avg_kernel_area_ratio', 'avg_width', 'avg_height' ], tablefmt = 'orgtbl')
    
    print(table + "\n")
    
    

    
    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')
        trait_file_csv = (args['result'] + 'trait.csv')
    else:
        trait_file = (file_path + 'trait.xlsx')
        trait_file_csv = (file_path + 'trait.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
        #sheet_leaf = wb.create_sheet()

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'tag_info'
        sheet.cell(row = 1, column = 3).value = 'avg_kernel_area'
        sheet.cell(row = 1, column = 4).value = 'avg_kernel_area_ratio'
        sheet.cell(row = 1, column = 5).value = 'avg_width'
        sheet.cell(row = 1, column = 6).value = 'avg_height'
        
        '''
        sheet.cell(row = 1, column = 6).value = 'color distribution cluster 1'
        sheet.cell(row = 1, column = 7).value = 'color distribution cluster 2'
        sheet.cell(row = 1, column = 8).value = 'color distribution cluster 3'
        sheet.cell(row = 1, column = 9).value = 'color distribution cluster 4'
        sheet.cell(row = 1, column = 10).value = 'color cluster 1 hex value'
        sheet.cell(row = 1, column = 11).value = 'color cluster 2 hex value'
        sheet.cell(row = 1, column = 12).value = 'color cluster 3 hex value'
        sheet.cell(row = 1, column = 13).value = 'color cluster 4 hex value'        
        '''
    
       
        
    for row in result_list:
        sheet.append(row)
   
    #for row in result_list_leaf:
        #sheet_leaf.append(row)
        
    #for row in result_list_leaf:
        #sheet_leaf.append(row)
    
    #save the csv file
    wb.save(trait_file)
    
    
    
    '''
    wb = openpyxl.load_workbook(trait_file)
    sh = wb.active # was .get_active_sheet()
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows: # generator; was sh.rows
            c.writerow([cell.value for cell in r])
    '''

    

    
