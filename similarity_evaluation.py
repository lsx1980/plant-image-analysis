"""
Version: 1.5

Summary: Compute the similarity of a pair of images

Author: suxing liu

Author-email: suxingliu@gmail.com

USAGE:

    python3 similarity_evaluation.py -o /org_img_path/ -p /pred_img_path/
    
"""

import subprocess, os
import sys
import argparse
import numpy as np 
import pathlib
import os
import glob
import shutil
from pathlib import Path
from scipy.spatial import KDTree
import cv2
import imutils

from image_similarity_measures.evaluate import evaluation

import openpyxl



def similarity_evaluation(image_ori, image_pred):
    
    file_name = Path(image_ori).name
            
            
    # metrics = ["rmse", "ssim"]
    dict_val = evaluation(org_img_path = image_ori, pred_img_path = image_pred, metrics = ["ssim"])
    
    
    values = [float(x) for x in list(dict_val.values())]
    


    return file_name, values[0]


        

if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--org_img_path", required = True, help = "path to original image file")
    ap.add_argument("-p", "--pred_img_path", required = True, help = "path to predicted image file")
    ap.add_argument("-ft", "--filetype", required = False,  default = "png", help = "Image filetype")
    args = vars(ap.parse_args())
    
    
    # setting path to model file
    org_file_path = args["org_img_path"]
    pred_file_path = args["pred_img_path"]
    ext = args['filetype']
    
    #accquire image file list
    filetype = '*.' + ext
    org_image_path = org_file_path + filetype
    pred_image_path = pred_file_path + filetype
    
    #accquire image file list
    org_imgList = sorted(glob.glob(org_image_path))
    pred_imgList = sorted(glob.glob(pred_image_path))
    

    
    #global save_path
    
    n_images = len(org_imgList)
    
    result_list = []

    file_name_list = []
    
    
        
    #loop execute
    
    for i, (image_ori, image_pred) in enumerate(zip(org_imgList, pred_imgList)):
        
        
        (file_name, values) = similarity_evaluation(image_ori, image_pred)
        
        
        result_list.append([file_name, values])

    
    #print(result_list)
    
    #print(type(result_list[0]))

    
    #########################################################################
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
   

    
    trait_file = (org_file_path + '/ssim_value.xlsx')

    
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
        

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        #sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'ssim'

        
    for row in result_list:
        sheet.append(row)
   

    #save the csv file
    wb.save(trait_file)
    
    
