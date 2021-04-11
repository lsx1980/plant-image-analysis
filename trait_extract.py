'''
Name: color_segmentation.py

Version: 1.0

Summary:  Extract plant traits (leaf area, width, height, ) by paralell processing 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2018-09-29

USAGE:

time python3 trait_extract.py -i 01.jpg


'''

# import the necessary packages
import os
import utils

import argparse
from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import watershed, medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage

import numpy as np
import argparse
import cv2

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches


from openpyxl import load_workbook
from openpyxl import Workbook

import warnings
warnings.filterwarnings("ignore")



MBFACTOR = float(1<<20)

class ComputeCurvature:

    def __init__(self,x,y):
        """ Initialize some variables """
        self.xc = 0  # X-coordinate of circle center
        self.yc = 0  # Y-coordinate of circle center
        self.r = 0   # Radius of the circle
        self.xx = np.array([])  # Data points
        self.yy = np.array([])  # Data points
        self.x = x  # X-coordinate of circle center
        self.y = y  # Y-coordinate of circle center

    def calc_r(self, xc, yc):
        """ calculate the distance of each 2D points from the center (xc, yc) """
        return np.sqrt((self.xx-xc)**2 + (self.yy-yc)**2)

    def f(self, c):
        """ calculate the algebraic distance between the data points and the mean circle centered at c=(xc, yc) """
        ri = self.calc_r(*c)
        return ri - ri.mean()

    def df(self, c):
        """ Jacobian of f_2b
        The axis corresponding to derivatives must be coherent with the col_deriv option of leastsq"""
        xc, yc = c
        df_dc = np.empty((len(c), self.x.size))

        ri = self.calc_r(xc, yc)
        df_dc[0] = (xc - self.x)/ri                   # dR/dxc
        df_dc[1] = (yc - self.y)/ri                   # dR/dyc
        df_dc = df_dc - df_dc.mean(axis=1)[:, np.newaxis]
        return df_dc

    def fit(self, xx, yy):
        self.xx = xx
        self.yy = yy
        center_estimate = np.r_[np.mean(xx), np.mean(yy)]
        center = optimize.leastsq(self.f, center_estimate, Dfun=self.df, col_deriv=True)[0]

        self.xc, self.yc = center
        ri = self.calc_r(*center)
        self.r = ri.mean()

        return 1 / self.r  # Return the curvature


# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        print (path + ' folder constructed!')
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        print (path+' path exists!')
        return False
        

def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (width, height, n_channel) = image.shape
    
    #print("image shape: \n")
    #print(width, height, n_channel)
    
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    

    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    #print("thresh.dtype")
    #print(thresh.dtype)
    
    
    
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh, connectivity=8)
    
    sizes = stats[1:, -1]
    
    nb_components = nb_components - 1
    
    min_size = 150 
    
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    #for every component in the image, you keep it only if it's above min_size
    for i in range(0, nb_components):
        if sizes[i] >= min_size:
            img_thresh[output == i + 1] = 255
    
    #from skimage import img_as_ubyte
    
    #img_thresh = img_as_ubyte(img_thresh)
    
    print("img_thresh.dtype")
    print(img_thresh.dtype)
    
    return img_thresh
    

def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis



def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    labels = watershed(-D, markers, mask = thresh)
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels


def comp_external_contour(orig,thresh):
    
    img_height, img_width, img_channels = orig.shape
    
    if args['color_space'] == 'gray':
        
        orig = np.zeros([img_height, img_width, 3], dtype=np.uint8)
    
    #find contours and get the external one
    #image_result, contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    contours = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
   
    #print("len(contours)")
    #print(len(contours))
    
    index = 1
    
    #trait_img = orig
    #area = orig
    #solidity = orig
    
    for c in contours:
        
        #get the bounding rect
        x, y, w, h = cv2.boundingRect(c)
        
        if w>img_width*0.5 and h>img_height*0.5:
            
            trait_img = cv2.drawContours(orig, contours, -1, (255, 255, 0), 1)
    
            # draw a green rectangle to visualize the bounding rect
            roi = orig[y:y+h, x:x+w]
            
            print("ROI {} detected ...\n".format(index))
            #result_file = (save_path +  str(index) + file_extension)
            #cv2.imwrite(result_file, roi)
            
            trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 3)
            
            index+= 1

            '''
            #get the min area rect
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            # convert all coordinates floating point values to int
            box = np.int0(box)
            #draw a red 'nghien' rectangle
            trait_img = cv2.drawContours(orig, [box], 0, (0, 0, 255))
            '''
             # get convex hull
            hull = cv2.convexHull(c)
            # draw it in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 3)
            
            '''
            # calculate epsilon base on contour's perimeter
            # contour's perimeter is returned by cv2.arcLength
            epsilon = 0.01 * cv2.arcLength(c, True)
            # get approx polygons
            approx = cv2.approxPolyDP(c, epsilon, True)
            # draw approx polygons
            trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
         
            # hull is convex shape as a polygon
            hull = cv2.convexHull(c)
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
            '''
            
            '''
            #get the min enclosing circle
            (x, y), radius = cv2.minEnclosingCircle(c)
            # convert all values to int
            center = (int(x), int(y))
            radius = int(radius)
            # and draw the circle in blue
            trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
            '''
            
            area = cv2.contourArea(c)
            print("Leaf area = {0:.2f}... \n".format(area))
            
            
            hull = cv2.convexHull(c)
            hull_area = cv2.contourArea(hull)
            solidity = float(area)/hull_area
            print("solidity = {0:.2f}... \n".format(solidity))
            
            extLeft = tuple(c[c[:,:,0].argmin()][0])
            extRight = tuple(c[c[:,:,0].argmax()][0])
            extTop = tuple(c[c[:,:,1].argmin()][0])
            extBot = tuple(c[c[:,:,1].argmax()][0])
            
            trait_img = cv2.circle(orig, extLeft, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extRight, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extTop, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extBot, 3, (255, 0, 0), -1)
            
            max_width = dist.euclidean(extLeft, extRight)
            max_height = dist.euclidean(extTop, extBot)
            
            if max_width > max_height:
                trait_img = cv2.line(orig, extLeft, extRight, (0,255,0), 2)
            else:
                trait_img = cv2.line(orig, extTop, extBot, (0,255,0), 2)
            
            print("Width and height are {0:.2f},{1:.2f}... \n".format(w, h))

    return trait_img, area, solidity, w, h
    
    

def compute_curv(orig, labels):
    
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    curv_sum = 0.0
    count = 0
    # curvature computation
    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start = 1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype = "uint8")
        mask[labels == label] = 255
     
        # detect contours in the mask and grab the largest one
        contours = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        #img, contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        
        c = max(contours, key = cv2.contourArea)
        
        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        label_trait = cv2.putText(orig, "#{}".format(label), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        #cv2.putText(orig, "#{}".format(curvature), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        
        
        if len(c) >= 5 :
            label_trait = cv2.drawContours(orig, [c], -1, (255, 0, 0), 2)
            ellipse = cv2.fitEllipse(c)
            label_trait = cv2.ellipse(orig,ellipse,(0,255,0),2)
            
            c_np = np.vstack(c).squeeze()
            count+=1
            
            x = c_np[:,0]
            y = c_np[:,1]
            
            comp_curv = ComputeCurvature(x, y)
            curvature = comp_curv.fit(x, y)
            
            curv_sum = curv_sum + curvature

        else:
            # optional to "delete" the small contours
            label_trait = cv2.drawContours(orig, [c], -1, (0, 0, 255), 2)
            print("lack of enough points to fit ellipse")
    
    
    print('average curvature = {0:.2f}'.format(curv_sum/count))
    
    return curv_sum/count, label_trait


def color_quantization(image, mask, save_path, num_clusters):
    
    #grab image width and height
    (h, w) = image.shape[:2]
    
    #change the color storage order
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    #apply the mask to get the segmentation of plant
    masked_image = cv2.bitwise_and(image, image, mask = mask)
       
    # reshape the image to be a list of pixels
    pixels = masked_image.reshape((masked_image.shape[0] * masked_image.shape[1], 3))
        
    ############################################################
    #Clustering process
    ###############################################################
    # cluster the pixel intensities
    clt = MiniBatchKMeans(n_clusters = num_clusters)
    #clt = KMeans(n_clusters = args["clusters"])
    clt.fit(pixels)

    #assign labels to each cluster 
    labels = clt.fit_predict(pixels)

    #obtain the quantized clusters using each label
    quant = clt.cluster_centers_.astype("uint8")[labels]

    # reshape the feature vectors to images
    quant = quant.reshape((h, w, 3))
    image_rec = pixels.reshape((h, w, 3))
    
    # convert from L*a*b* to RGB
    quant = cv2.cvtColor(quant, cv2.COLOR_RGB2BGR)
    image_rec = cv2.cvtColor(image_rec, cv2.COLOR_RGB2BGR)
    
    # display the images and wait for a keypress
    #cv2.imshow("image", np.hstack([image_rec, quant]))
    #cv2.waitKey(0)
    
    #define result path for labeled images
    result_img_path = save_path + 'cluster_out.png'
    
    # save color_quantization results
    cv2.imwrite(result_img_path,quant)

    # build a histogram of clusters and then create a figure representing the number of pixels labeled to each color
    hist = utils.centroid_histogram(clt)

    # remove the background color cluster
    clt.cluster_centers_ = clt.cluster_centers_[1: len(clt.cluster_centers_)]

    #build a histogram of clusters using center lables
    numLabels = utils.plot_centroid_histogram(save_path,clt)

    #create a figure representing the distribution of each color
    bar = utils.plot_colors(hist, clt.cluster_centers_)

    #save a figure of color bar 
    utils.plot_color_bar(save_path, bar)


def extract_traits(image):
    
    #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    abs_path = os.path.abspath(args['image'])
    
    filename, file_extension = os.path.splitext(args['image'])
    
    file_size = os.path.getsize(args['image'])/MBFACTOR
    
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    base_name = os.path.splitext(os.path.basename(filename))[0]
    # save folder construction
    mkpath = os.path.dirname(abs_path) +'/' + base_name
    mkdir(mkpath)
    save_path = mkpath + '/'
    print ("results_folder: " + save_path)
    
    if (file_size > 5.0):
        print("It will take some time due to larger file size {0} MB".format(str(int(file_size))))
    else:
        print("Segmentaing plant object using automatic color clustering method... ")
    
    #make backup image
    orig = image.copy()
     
    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']
    
    #color clustering based plant object segmentation
    thresh = color_cluster_seg(orig, args_colorspace, args_channels, args_num_clusters)
    
    # save segmentation result
    result_file = (save_path + filename + '_seg' + file_extension)
    #print(filename)
    cv2.imwrite(result_file, thresh)
    
    num_clusters = 5
    #save color quantization result
    color_quantization(image, thresh, save_path, num_clusters)
    
    
    #accquire medial axis of segmentation mask
    image_medial_axis = medial_axis_image(thresh)

    # save medial axis result
    result_file = (save_path + filename + '_medial_axis' + file_extension)
    cv2.imwrite(result_file, img_as_ubyte(image_medial_axis))
    
    
    min_distance_value = 5
    #watershed based leaf area segmentaiton 
    labels = watershed_seg(orig, thresh, min_distance_value)

    #save watershed result label image
     #Map component labels to hue val
    label_hue = np.uint8(128*labels/np.max(labels))

    blank_ch = 255*np.ones_like(label_hue)
    labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

    # cvt to BGR for display
    labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

    # set background label to black
    labeled_img[label_hue==0] = 0
    result_file = (save_path + filename + '_label' + file_extension)
    #plt.imsave(result_file, img_as_float(labels), cmap = "Spectral")
    cv2.imwrite(result_file, labeled_img)
    
    (avg_curv, label_trait) = compute_curv(orig, labels)
    
     #save watershed result label image
    result_file = (save_path + filename + '_curv' + file_extension)
    cv2.imwrite(result_file, label_trait)
    
    
    #find external contour 
    (trait_img, area, solidity, max_width, max_height) = comp_external_contour(image.copy(),thresh)
    # save segmentation result
    result_file = (save_path + filename + '_excontour' + file_extension)
    #print(filename)
    cv2.imwrite(result_file, trait_img)   
    
    
    return filename,area, solidity, max_width, max_height, avg_curv


if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument('-i', '--image', required = True, help = 'Path to image file')
    ap.add_argument('-s', '--color-space', type =str, default ='lab', help='Color space to use: BGR (default), HSV, Lab, YCrCb (YCC)')
      args = vars(ap.parse_args())
    

    # load original image
    image = cv2.imread(args['image'])
    
    print(image.shape)
    print(image.dtype)
    
    if args['color_space'] == 'gray':
        
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        image = cv2.cvtColor(gray,cv2.COLOR_GRAY2RGB)
    
    #extract_traits(image)
    
    
    (base_name, area, solidity, max_width, max_height, avg_curv) = extract_traits(image)
    
    
    # Write data
    trait_file = (os.path.dirname(os.path.abspath(args['image'])) +'/'  + 'trait.xlsx')
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.get_active_sheet()

    else:
        # Keep presets
        wb = Workbook()
        sheet = wb.active

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'leaf_area'
        sheet.cell(row = 1, column = 3).value = 'solidity'
        sheet.cell(row = 1, column = 4).value = 'max_width'
        sheet.cell(row = 1, column = 5).value = 'max_height'
        sheet.cell(row = 1, column = 6).value = 'curvature'
    
    
    sheet.cell(row = int(base_name) + 1, column = 1).value = base_name 
    sheet.cell(row = int(base_name) + 1, column = 2).value = area 
    sheet.cell(row = int(base_name) + 1, column = 3).value = solidity 
    sheet.cell(row = int(base_name) + 1, column = 4).value = max_width 
    sheet.cell(row = int(base_name) + 1, column = 5).value = max_height 
    sheet.cell(row = int(base_name) + 1, column = 6).value = avg_curv 
    
    #save the csv file
    wb.save(trait_file)
    
    
    

    
    

    

    
