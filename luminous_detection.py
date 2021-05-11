'''
Name: luminous_detection.py

Version: 1.0

Summary: Detect dark images by converting it to LAB color space to access the luminous channel which is independent of colors.

         Add crop image based on marker location and image enhancement modules
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2021-04-09

USAGE:

time python3 luminous_detection.py -p ~/plant-image-analysis/test/ -ft png 

'''

# import necessary packages
import argparse
import cv2
import numpy as np
import os
import glob
from pathlib import Path 

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from tabulate import tabulate
import openpyxl
from PIL import Image, ImageEnhance


import itertools

# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False

# get base name of a file from full path
def get_basename(image_file):
    
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    
    base_name = os.path.splitext(os.path.basename(filename))[0]

    return base_name
    
    #image_file_name = Path(image_file).name


# Convert it to LAB color space to access the luminous channel which is independent of colors.
def isbright(image_file):
    
    # Set up threshold value for luminous channel, can be adjusted and generalized 
    thresh = 0.5
    
    # Load image file 
    orig = cv2.imread(image_file)
    
    # Make backup image
    image = orig.copy()
    
    # Get file name
    #abs_path = os.path.abspath(image_file)
    
    #filename, file_extension = os.path.splitext(abs_path)
    #base_name = os.path.splitext(os.path.basename(filename))[0]

    image_file_name = Path(image_file).name
    
    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    
    # Normalize L channel by dividing all pixel values with maximum pixel value
    L = L/np.max(L)
    
    text_bool = "bright" if np.mean(L) < thresh else "dark"
    
    return image_file_name, np.mean(L), text_bool
    


# get weight value based on liner interpolation
def blend_weight_calculator(left_image_idx, right_image_idx, current_image_idx):
    
    window_width = right_image_idx - left_image_idx
    
    if window_width > 0:
        left_weight = abs(right_image_idx - current_image_idx)/window_width
                    
        right_weight = abs(current_image_idx - left_image_idx)/window_width
    else:
        left_weight = 0.5
        right_weight = 0.5
    
    return left_weight,right_weight


# blend two images based on weights
def blend_image(left_image, right_image, left_weight, right_weight):
    
    left_img = cv2.imread(left_image)
    
    right_img = cv2.imread(right_image)
    
    blended = cv2.addWeighted(left_img, left_weight, right_img, right_weight, 0)
    
    return blended


# detect dark image and replac them with liner interpolated image
def check_discard_merge(imgList):
    
    # create and assign index list for dark image
    idx_dark_imglist = [0] * len(imgList)
    
    result_list = []
    
    
    for idx, image in enumerate(imgList):
        
        img_name, mean_luminosity, luminosity_str = isbright(image)  # luminosity detection, luminosity_str is either 'dark' or 'bright'
        
        result_list.append([img_name, mean_luminosity, luminosity_str])
        
        idx_dark_imglist[idx] = -1 if luminosity_str == 'dark' else (idx)
    
    # Output sum table in command window 
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    table = tabulate(result_list, headers = ['image_file_name', 'luminous_avg', 'dark_or_bright'], tablefmt = 'orgtbl')

    print(table + "\n")
    
    # save dark image detection result as excel file
    result_excel(result_list, file_path)

    #print(idx_dark_imglist)
    
    #Finding consecutive occurrences of -1 in an array
    max_dark_list_length = max(len(list(v)) for g,v in itertools.groupby(idx_dark_imglist))
    
    #check max dark image sequence length, current method only deal with case with length equals 2
    #print(max_dark_list_length)
    
    #find dark image index 
    idx_dark = [i for i,x in enumerate(idx_dark_imglist) if x == -1]
    
    #print(idx_dark)
    
    #print(len(idx_dark_imglist))
    
    # process dark image 
    if len(idx_dark) > 1:
        
        for idx, value in enumerate(idx_dark):
            
            #print("current value:{0}".format(value))
            
            # if dark image appears as the start of image list
            if value == 0:

                right_image_idx = ((value+1) if idx_dark_imglist[value+1] != -1 else (value+2))
                
                left_image_idx = right_image_idx
            
            # if dark image appears as the end of image list
            elif value == len(idx_dark_imglist)-1:
                
                left_image_idx = ((value-1) if idx_dark_imglist[value-1] != -1 else (value-2))
                
                right_image_idx = left_image_idx
            
            else:
                
                left_image_idx = ((value-1) if idx_dark_imglist[value-1] != -1 else (value-2))
                
                right_image_idx = ((value+1) if idx_dark_imglist[value+1] != -1 else (value+2))

            
            #print("current image idx:{0}, left_idx:{1}, right_idx:{2}\n".format(value, left_image_idx, right_image_idx))
            
            (left_weight, right_weight) = blend_weight_calculator(left_image_idx, right_image_idx, value)
            
            #print("current image idx:{0}, left_idx:{1}, right_idx:{2}, left_weight:{3}, right_weight:{4} \n".format(value, left_image_idx, right_image_idx, left_weight, right_weight))
            

            blended = blend_image(imgList[left_image_idx], imgList[right_image_idx], left_weight, right_weight)
 
            print("blending current image:{0}, left:{1}, right:{2}, left_weight:{3:.2f}, right_weight:{4:.2f} \n".format(get_basename(imgList[value]), get_basename(imgList[left_image_idx]), get_basename(imgList[right_image_idx]), left_weight, right_weight))
    
            # save result by overwriting original files
            result_file = (imgList[value])
            cv2.imwrite(result_file, blended)
            
            #save result into result folder for debugging
            result_file = (save_path + get_basename(imgList[value]) + '.' + args['filetype'])
            cv2.imwrite(result_file, blended)


# Detect circles in the image
def circle_detect(image_file):
   
    print("Cropping image {}...\n".format(image_file))
    
    # load the image, clone it for output, and then convert it to grayscale
    img_ori = cv2.imread(image_file)
    
    img_rgb = img_ori.copy()
      
    # Convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED) 
    
    # Specify a threshold 
    threshold = 0.8
      
    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= threshold)  
    
    if len(loc):
    
        (y,x) = np.unravel_index(res.argmax(), res.shape)
    
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)
    
        #print(y,x)
        
        #print(min_val, max_val, min_loc, max_loc)
    
        # Draw a rectangle around the matched region. 
        for pt in zip(*loc[::-1]): 
            circle_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,255), 2) 

        # save segmentation result
        #result_file = (save_path + base_name + '_circle.' + args['filetype'])
        #print(result_file)
        #cv2.imwrite(result_file, circle_overlay)
        
        crop_img = img_rgb[y+100:y+800, x-700:x]
        
        # save segmentation result
        #result_file = (save_path + base_name + '_cropped.' + args['filetype'])
        #print(result_file)
        #cv2.imwrite(result_file, crop_img)

    return crop_img
    

def image_enhance(image_file):

    im = Image.fromarray(cv2.cvtColor(image_file, cv2.COLOR_BGR2RGB))
    
    im_sharpness = ImageEnhance.Sharpness(im).enhance(3.5)
    
    im_contrast = ImageEnhance.Contrast(im_sharpness).enhance(1.5)

    im_out = ImageEnhance.Brightness(im_contrast).enhance(1.2)
    
    return im_out



def result_excel(result_list, file_path):
    
    result_file = (file_path + '/luminous_detection.xlsx')
    
    #print(result_file)
    
    if os.path.isfile(result_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(result_file)

        #Get the current Active Sheet
        sheet = wb.active

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.cell(row = 1, column = 1).value = 'image_file_name'
        sheet.cell(row = 1, column = 2).value = 'luminous_avg'
        sheet.cell(row = 1, column = 3).value = 'dark_or_bright'
    
    for row in result_list:
        sheet.append(row)
    
    #save the csv file
    wb.save(result_file)
    

if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required=True,    help = "image filetype")

    args = vars(ap.parse_args())
    
    
    # Setting path to image files
    file_path = args["path"]
    ext = args['filetype']

    # Extract file type and path
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    # Accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    global save_path, template
   
    mkpath = os.path.dirname(file_path) +'/merged'
    mkdir(mkpath)
    save_path = mkpath + '/'
    
    template_path = "/home/suxing/plant-image-analysis/marker_template/template.png"
    # Read the template 
    template = cv2.imread(template_path, 0) 
    
    #print((imgList))
    #global save_path
    
    # Get number of images in the data folder
    n_images = len(imgList)
    
    # replace dark image using blended image
    #check_discard_merge(imgList)
    
    # enhance image Contrast,Brightness,Sharpness
    for image in imgList:
       
        im_out_name = get_basename(image)
        
        # construct the result file path
        result_img_path = save_path + im_out_name + '.' + ext
        
        crop_img = circle_detect(image)
        
        im_out = image_enhance(crop_img)

        im_out.save(result_img_path)
    
    
    '''
    # Loop execute
    for image in imgList:
        
        (image_file_name, luminous_avg, dark_or_bright) = isbright(image)
        
        result_list.append([image_file_name, luminous_avg, dark_or_bright])
    '''
    '''
    # Parallel processing
    
    # get cpu number for parallel processing
    agents = psutil.cpu_count()   
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result_list = pool.map(isbright, imgList)
        pool.terminate()
    

    # Output sum table in command window 
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    table = tabulate(result_list, headers = ['image_file_name', 'luminous_avg', 'dark_or_bright'], tablefmt = 'orgtbl')

    print(table + "\n")
    
    result_excel(result_list, file_path)
    '''
