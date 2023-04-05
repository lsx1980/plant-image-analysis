'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract plant shoot traits (larea, solidity, max_width, max_height, avg_curv, color_cluster) by paralell processing 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2018-09-29

USAGE:

time python3 trait_extract_parallel_tara_root.py -p ~/example/Tara_data/test/ -ft jpg -s YCC -c 2 -min 10000

'''

# import the necessary packages
import os
import glob
import utils

from collections import Counter
from collections import OrderedDict

import argparse

from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.color import rgb2lab, deltaE_cie76
from skimage import morphology
from skimage.segmentation import clear_border, watershed
from skimage.measure import regionprops

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage
from scipy.interpolate import interp1d

#from skan import skeleton_to_csgraph, Skeleton, summarize, draw

#import networkx as nx

import imutils

import numpy as np
import argparse
import cv2

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import math
import openpyxl
import csv
    
from tabulate import tabulate

import warnings
warnings.filterwarnings("ignore")

import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

from pathlib import Path


from matplotlib import collections

import matplotlib.colors



MBFACTOR = float(1<<20)

class ComputeCurvature:

    def __init__(self,x,y):
        """ Initialize some variables """
        self.xc = 0  # X-coordinate of circle center
        self.yc = 0  # Y-coordinate of circle center
        self.r = 0   # Radius of the circle
        self.xx = np.array([])  # Data points
        self.yy = np.array([])  # Data points
        self.x = x  # X-coordinate of circle center
        self.y = y  # Y-coordinate of circle center

    def calc_r(self, xc, yc):
        """ calculate the distance of each 2D points from the center (xc, yc) """
        return np.sqrt((self.xx-xc)**2 + (self.yy-yc)**2)

    def f(self, c):
        """ calculate the algebraic distance between the data points and the mean circle centered at c=(xc, yc) """
        ri = self.calc_r(*c)
        return ri - ri.mean()

    def df(self, c):
        """ Jacobian of f_2b
        The axis corresponding to derivatives must be coherent with the col_deriv option of leastsq"""
        xc, yc = c
        df_dc = np.empty((len(c), self.x.size))

        ri = self.calc_r(xc, yc)
        df_dc[0] = (xc - self.x)/ri                   # dR/dxc
        df_dc[1] = (yc - self.y)/ri                   # dR/dyc
        df_dc = df_dc - df_dc.mean(axis=1)[:, np.newaxis]
        return df_dc

    def fit(self, xx, yy):
        self.xx = xx
        self.yy = yy
        center_estimate = np.r_[np.mean(xx), np.mean(yy)]
        center = optimize.leastsq(self.f, center_estimate, Dfun=self.df, col_deriv=True)[0]

        self.xc, self.yc = center
        ri = self.calc_r(*center)
        self.r = ri.mean()

        return 1 / self.r  # Return the curvature



class ColorLabeler:
    def __init__(self):
        # initialize the colors dictionary, containing the color
        # name as the key and the RGB tuple as the value
        colors = OrderedDict({
            "dark skin": (115, 82, 68),
            "light skin": (194, 150, 130),
            "blue sky": (98, 122, 157),
            "foliage": (87, 108, 67),
            "blue flower": (133, 128, 177),
            "bluish green": (103, 189, 170),
            "orange": (214, 126, 44),
            "purplish blue": (8, 91, 166),
            "moderate red": (193, 90, 99),
            "purple": (94, 60, 108),
            "yellow green": (157, 188, 64),
            "orange yellow": (224, 163, 46),
            "blue": (56, 61, 150),
            "green": (70, 148, 73),
            "red": (175, 54, 60),
            "yellow": (231, 199, 31),
            "magneta": (187, 86, 149),
            "cyan": (8, 133, 161),
            "white": (243, 243, 242),
            "neutral 8": (200, 200, 200),
            "neutral 6.5": (160, 160, 160),
            "neutral 5": (122, 122, 121),
            "neutral 3.5": (85, 85, 85),
            "black": (52, 52, 52)})
        # allocate memory for the L*a*b* image, then initialize
        # the color names list
        self.lab = np.zeros((len(colors), 1, 3), dtype="uint8")
        self.colorNames = []
        # loop over the colors dictionary
        for (i, (name, rgb)) in enumerate(colors.items()):
            # update the L*a*b* array and the color names list
            self.lab[i] = rgb
            self.colorNames.append(name)
        # convert the L*a*b* array from the RGB color space
        # to L*a*b*
        self.lab = cv2.cvtColor(self.lab, cv2.COLOR_RGB2LAB)

    def label(self, image, c):
            # construct a mask for the contour, then compute the
            # average L*a*b* value for the masked region
            mask = np.zeros(image.shape[:2], dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)
            mask = cv2.erode(mask, None, iterations=2)
            mean = cv2.mean(image, mask=mask)[:3]
            # initialize the minimum distance found thus far
            minDist = (np.inf, None)
            # loop over the known L*a*b* color values
            for (i, row) in enumerate(self.lab):
                # compute the distance between the current L*a*b*
                # color value and the mean of the image
                d = dist.euclidean(row[0], mean)
                # if the distance is smaller than the current distance,
                # then update the bookkeeping variable
                if d < minDist[0]:
                    minDist = (d, i)
            # return the name of the color with the smallest distance
            return self.colorNames[minDist[1]]




# generate foloder to store the output results
def mkdir(path):
    # import module
    import os
 
    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False
        

def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
            
    (width, height, n_channel) = image.shape
    
    #print("image shape: \n")
    #print(width, height, n_channel)
    
 
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    

    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    #thresh_cleaned = (thresh)
    
    
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned = clear_border(thresh)
    else:
        thresh_cleaned = thresh
    
     
    (nb_components, output, stats, centroids) = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)

    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    sizes = stats[1:, cv2.CC_STAT_AREA]
    
    Coord_left = stats[1:, cv2.CC_STAT_LEFT]
    
    Coord_top = stats[1:, cv2.CC_STAT_TOP]
    
    Coord_width = stats[1:, cv2.CC_STAT_WIDTH]
    
    Coord_height = stats[1:, cv2.CC_STAT_HEIGHT]
    
    Coord_centroids = centroids
    
    #print("Coord_centroids {}\n".format(centroids[1][1]))
    
    #print("[width, height] {} {}\n".format(width, height))
    
    nb_components = nb_components - 1
    
    
    
    max_size = width*height*0.1
    
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    #for every component in the image, keep it only if it's above min_size
    for i in range(0, nb_components):
        
        '''
        #print("{} nb_components found".format(i))
        
        if (sizes[i] >= min_size) and (Coord_left[i] > 1) and (Coord_top[i] > 1) and (Coord_width[i] - Coord_left[i] > 0) and (Coord_height[i] - Coord_top[i] > 0) and (centroids[i][0] - width*0.5 < 10) and ((centroids[i][1] - height*0.5 < 10)) and ((sizes[i] <= max_size)):
            img_thresh[output == i + 1] = 255
            
            print("Foreground center found ")
            
        elif ((Coord_width[i] - Coord_left[i])*0.5 - width < 15) and (centroids[i][0] - width*0.5 < 15) and (centroids[i][1] - height*0.5 < 15) and ((sizes[i] <= max_size)):
            imax = max(enumerate(sizes), key=(lambda x: x[1]))[0] + 1    
            img_thresh[output == imax] = 255
            print("Foreground max found ")
        '''
        
        if (sizes[i] >= min_size):
        
            img_thresh[output == i + 1] = 255
        
    #from skimage import img_as_ubyte
    
    #img_thresh = img_as_ubyte(img_thresh)
    
    #print("img_thresh.dtype")
    #print(img_thresh.dtype)
    
    
    size_kernel = 1
    
    #if mask contains mutiple non-conected parts, combine them into one. 
    contours, hier = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) > 1:
        
        print("mask contains mutiple non-conected parts, combine them into one\n")
        
        kernel = np.ones((size_kernel,size_kernel), np.uint8)

        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        img_thresh = closing
        
    
    
    
    #return img_thresh
    return img_thresh
    
    #return thresh_cleaned
    

def medial_axis_image(thresh):
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))
    
    image_medial_axis = medial_axis(image_bw)
    
    return image_medial_axis


def skeleton_bw(thresh):

    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it
    
    #convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    #skeleton = morphology.skeletonize(image_bw)
    
    skeleton = morphology.thin(image_bw)

    skeleton_img = skeleton.astype(np.uint8) * 255

    return skeleton_img, skeleton


def watershed_seg(orig, thresh, min_distance_value):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(localMax, structure = np.ones((3, 3)))[0]
    
    #print("markers")
    #print(type(markers))
    
    labels = watershed(-D, markers, mask = thresh)
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels


def percentage(part, whole):
  
  #percentage = "{:.0%}".format(float(part)/float(whole))
  
  percentage = "{:.2f}".format(float(part)/float(whole))
  
  return str(percentage)



    
    

def circle_detection(image):

    """Detecting Circles in Images using OpenCV and Hough Circles
    
    Inputs: 
    
        image: image loaded 

    Returns:
    
        circles: detcted circles
        
        circle_detection_img: circle overlayed with image
        
        diameter_circle: diameter of detected circle
        
    """
    
    # create background image for drawing the detected circle
    output = image.copy()
    
    # obtain image dimension
    img_height, img_width, n_channels = image.shape
    
    #backup input image
    circle_detection_img = image.copy()
    
    # change image from RGB to Gray scale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # apply blur filter 
    blurred = cv2.medianBlur(gray, 25)
    
    # setup parameters for circle detection
    
    # This parameter is the inverse ratio of the accumulator resolution to the image resolution 
    #(see Yuen et al. for more details). Essentially, the larger the dp gets, the smaller the accumulator array gets.
    dp = 1.5
    
    #Minimum distance between the center (x, y) coordinates of detected circles. 
    #If the minDist is too small, multiple circles in the same neighborhood as the original may be (falsely) detected. 
    #If the minDist is too large, then some circles may not be detected at all.
    minDist = 100
    
    #Gradient value used to handle edge detection in the Yuen et al. method.
    #param1 = 30
    
    #accumulator threshold value for the cv2.HOUGH_GRADIENT method. 
    #The smaller the threshold is, the more circles will be detected (including false circles). 
    #The larger the threshold is, the more circles will potentially be returned. 
    #param2 = 30  
    
    #Minimum/Maximum size of the radius (in pixels).
    #minRadius = 80
    #maxRadius = 120 
    
    # detect circles in the image
    #circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, 1.2, minDist, param1=param1, param2=param2, minRadius=minRadius, maxRadius=maxRadius)
    
    # detect circles in the image
    circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, dp, minDist)
    
    # initialize diameter of detected circle
    diameter_circle = 0
    
    
    circle_center_coord = []
    circle_center_radius = []
    idx_closest = 0
    
    
    # At leaset one circle is found
    if circles is not None:
        
        # Get the (x, y, r) as integers, convert the (x, y) coordinates and radius of the circles to integers
        circles = np.round(circles[0, :]).astype("int")
       
        if len(circles) < 2:
           
            print("Only one circle was found!\n")
           
        else:
            
            print("More than one circles were found!\n")
        
            idx_closest = 0
        
            #cv2.circle(output, (x, y), r, (0, 255, 0), 2)
          
        # loop over the circles and the (x, y) coordinates to get radius of the circles
        for (x, y, r) in circles:
            
            coord = (x, y)
            
            circle_center_coord.append(coord)
            circle_center_radius.append(r)

        if idx_closest == 0:

            print("Circle marker with radius = {} was detected!\n".format(circle_center_radius[idx_closest]))
        
        '''
        # draw the circle in the output image, then draw a center
        circle_detection_img = cv2.circle(output, circle_center_coord[idx_closest], circle_center_radius[idx_closest], (0, 255, 0), 4)
        circle_detection_img = cv2.circle(output, circle_center_coord[idx_closest], 5, (0, 128, 255), -1)

        # compute the diameter of coin
        diameter_circle = circle_center_radius[idx_closest]*2


        tmp_mask = np.zeros([img_width, img_height], dtype=np.uint8)

        tmp_mask = cv2.circle(tmp_mask, circle_center_coord[idx_closest], circle_center_radius[idx_closest] + 5, (255, 255, 255), -1)

        tmp_mask_binary = cv2.threshold(tmp_mask, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]

        masked_tmp = cv2.bitwise_and(image.copy(), image.copy(), mask = ~tmp_mask_binary)
        '''

        (startX, startY) = circle_center_coord[idx_closest]

        endX = startX + int(r*1.2) + 1050
        endY = startY + int(r*1.2) + 1050

        #sticker_crop_img = output[startY:endY, startX:endX]
        
        sticker_crop_img = output
    
    else:
        
        print("No circle was found!\n")
        
        sticker_crop_img = output
        
        diameter_circle = 0
    
    return circles, sticker_crop_img, diameter_circle



# Detect tray in the image
def tray_detect(img_ori, template):
    
    '''
    image_file_name = Path(image_file).name
    
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    print("Processing image : {0}\n".format(str(image_file)))
     
    # save folder construction
    mkpath = os.path.dirname(abs_path) +'/cropped'
    mkdir(mkpath)
    save_path = mkpath + '/'
    print ("results_folder: " + save_path)
    '''
   

    # load the image, clone it for output, and then convert it to grayscale
    img_rgb = img_ori.copy()
    
    # Convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # Store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    
    #(minVal, maxVal, minLoc, maxLoc) = cv2.minMaxLoc(res)
    
    
    # Specify a threshold 
    threshold = 0.6
    
    if np.amax(res) > threshold:
        
        flag = True
    else:

        flag = False
    
    print(flag)
    

    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= threshold)  
    
    if len(loc):
    
        (y,x) = np.unravel_index(res.argmax(), res.shape)
    
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)
    
        #print(y,x)
        
        #print(min_val, max_val, min_loc, max_loc)
        
        
        (startX, startY) = max_loc
        startX = startX + 0
        startY = startY + 50
        endX = startX + template.shape[0]
        endY = startY + template.shape[1]-100
        
        '''
        (startX, startY) = max_loc
        startX = startX - 100
        startY = startY
        
        endX = startX + template.shape[1] + int(w*0.8)
        endY = startY + template.shape[0] + int(h*0.8)
        '''
        '''
        # Draw a rectangle around the matched region. 
        for pt in zip(*loc[::-1]): 
            
            sticker_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,255), 2)
        '''
        
        sticker_crop_img = img_rgb[startY:endY, startX:endX]


    return  sticker_crop_img

'''
def individual_object_seg(orig, labels, save_path, base_name, file_extension):
    
    num_clusters = 5
    
    (width, height, n_channel) = orig.shape
    
    for label in np.unique(labels):
    # if the label is zero, we are examining the 'background'
    # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros((width, height), dtype="uint8")
        mask[labels == label] = 255
        
        # apply individual object mask
        masked = cv2.bitwise_and(orig, orig, mask = mask)
        
        # detect contours in the mask and grab the largest one
        #cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        #contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        #c = max(contours, key = cv2.contourArea)
        
        #if len(c) >= 5 :
            #label_img = cv2.drawContours(masked, [c], -1, (255, 0, 0), 2)
        
        mkpath_leaf = os.path.dirname(save_path) + '/leaf' + str(label)
        mkdir(mkpath_leaf)
        save_path_leaf = mkpath_leaf + '/'
        
        
        
        #define result path 
        result_img_path = (save_path_leaf + 'leaf_' + str(label) + file_extension)
        cv2.imwrite(result_img_path, masked)
        
        #save color quantization result
        #rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
        rgb_colors, counts = color_region(masked, mask, save_path_leaf, num_clusters)
        
        list_counts = list(counts.values())
        
        #print(type(list_counts))
        
        for value in list_counts:
            
            print(percentage(value, np.sum(list_counts)))
'''

        
        
    
    
        


'''
def watershed_seg_marker(orig, thresh, min_distance_value, img_marker):
    
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this
    # distance map
    D = ndimage.distance_transform_edt(thresh)
    
    gray = cv2.cvtColor(img_marker, cv2.COLOR_BGR2GRAY)
    img_marker = cv2.threshold(gray, 128, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
    
    #localMax = peak_local_max(D, indices = False, min_distance = min_distance_value,  labels = thresh)
     
    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then appy the Watershed algorithm
    markers = ndimage.label(img_marker, structure = np.ones((3, 3)))[0]

    labels = watershed(-D, markers, mask = thresh)

    
    props = regionprops(labels)
    
    areas = [p.area for p in props]
    
    import statistics
    
    
    #outlier_list = outlier_doubleMAD(areas, thresh = 1.0)
    
    #indices = [i for i, x in enumerate(outlier_list) if x]
    
    print(areas)
    print(statistics.mean(areas))
    #
    #print(outlier_list)
    #print(indices)
    
    
    print("[INFO] {} unique segments found\n".format(len(np.unique(labels)) - 1))
    
    return labels
'''


def comp_external_contour(orig,thresh):
    
    #find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
   
    img_height, img_width, img_channels = orig.shape
   
    index = 1
    
    trait_img = orig.copy()
    
    area = 0
    
    solidity = 0
    
    w=h=0
    
    for c in contours:
        
        #get the bounding rect
        x, y, w, h = cv2.boundingRect(c)
        
        if w>img_width*0.01 and h>img_height*0.01:
            
            trait_img = cv2.drawContours(orig, contours, -1, (255, 255, 0), 1)
    
            # draw a green rectangle to visualize the bounding rect
            roi = orig[y:y+h, x:x+w]
            
            print("ROI {} detected ...\n".format(index))
            #result_file = (save_path +  str(index) + file_extension)
            #cv2.imwrite(result_file, roi)
            
            trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 3)
            
            index+= 1

            '''
            #get the min area rect
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            # convert all coordinates floating point values to int
            box = np.int0(box)
            #draw a red 'nghien' rectangle
            trait_img = cv2.drawContours(orig, [box], 0, (0, 0, 255))
            '''
             # get convex hull
            hull = cv2.convexHull(c)
            # draw it in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 3)
            
            '''
            # calculate epsilon base on contour's perimeter
            # contour's perimeter is returned by cv2.arcLength
            epsilon = 0.01 * cv2.arcLength(c, True)
            # get approx polygons
            approx = cv2.approxPolyDP(c, epsilon, True)
            # draw approx polygons
            trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
         
            # hull is convex shape as a polygon
            hull = cv2.convexHull(c)
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
            '''
            
            '''
            #get the min enclosing circle
            (x, y), radius = cv2.minEnclosingCircle(c)
            # convert all values to int
            center = (int(x), int(y))
            radius = int(radius)
            # and draw the circle in blue
            trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
            '''
            
            area = cv2.contourArea(c)
            print("Leaf area = {0:.2f}... \n".format(area))
            
            
            hull = cv2.convexHull(c)
            hull_area = cv2.contourArea(hull)
            solidity = float(area)/hull_area
            print("solidity = {0:.2f}... \n".format(solidity))
            
            extLeft = tuple(c[c[:,:,0].argmin()][0])
            extRight = tuple(c[c[:,:,0].argmax()][0])
            extTop = tuple(c[c[:,:,1].argmin()][0])
            extBot = tuple(c[c[:,:,1].argmax()][0])
            
            trait_img = cv2.circle(orig, extLeft, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extRight, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extTop, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extBot, 3, (255, 0, 0), -1)
            
            max_width = dist.euclidean(extLeft, extRight)
            max_height = dist.euclidean(extTop, extBot)
            
            if max_width > max_height:
                trait_img = cv2.line(orig, extLeft, extRight, (0,255,0), 2)
            else:
                trait_img = cv2.line(orig, extTop, extBot, (0,255,0), 2)

            print("Width and height are {0:.2f},{1:.2f}... \n".format(w, h))
            
            
            
    return trait_img, area, solidity, w, h
    
    
# individual leaf object segmentation and traits computation
def leaf_traits_computation(orig, labels, save_path, base_name, file_extension):
    
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)
    
    leaf_index_rec = []
    contours_rec = []
    area_rec = []
    curv_rec = []
    solidity_rec = []
    major_axis_rec = []
    minor_axis_rec = []
    
    leaf_color_ratio_rec = []
    leaf_color_value_rec = []
    
    box_coord_rec = []
    
    count = 0
    
    num_clusters = 5
    

    # curvature computation
    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start = 1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue
     
        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype = "uint8")
        mask[labels == label] = 255
        
        
        #get the medial axis of the contour
        image_skeleton, skeleton = skeleton_bw(mask)

                
        # apply individual object mask
        masked = cv2.bitwise_and(orig, orig, mask = mask)
        
        #individual leaf segmentation and color analysis
        ################################################################################
        mkpath_leaf = os.path.dirname(save_path) + '/leaf' + str(label)
        mkdir(mkpath_leaf)
        save_path_leaf = mkpath_leaf + '/'
        

        #define result path 
        result_img_path = (save_path_leaf + 'leaf_' + str(label) + file_extension)
        cv2.imwrite(result_img_path, masked)
        
        # save skeleton result
        result_file = (save_path_leaf + 'leaf_skeleton_' + str(label) + file_extension)
        cv2.imwrite(result_file, img_as_ubyte(image_skeleton))
        
        
        #save color quantization result
        #rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
        (rgb_colors, counts, hex_colors) = color_region(masked, mask, save_path_leaf, num_clusters)
        
        #print("hex_colors = {} {}\n".format(hex_colors, type(hex_colors)))
        
        list_counts = list(counts.values())
        
        #list_hex_colors = list(hex_colors)
        
        #print(type(list_counts))
        
        color_ratio = []
        
        for value_counts, value_hex in zip(list_counts, hex_colors):
            
            #print(percentage(value, np.sum(list_counts)))
            
            color_ratio.append(percentage(value_counts, np.sum(list_counts)))
            
            #print("value_hex = {0}".format(value_hex))
            
            #value_hex.append(value_hex)
            
            
        leaf_color_ratio_rec.append(color_ratio)
        leaf_color_value_rec.append(hex_colors)
        
        # detect contours in the mask and grab the largest one
        #cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(),cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key = cv2.contourArea)
        
       
        if len(c) >= 10 :

            contours_rec.append(c)
            area_rec.append(cv2.contourArea(c))

        else:
            # optional to "delete" the small contours
            #label_trait = cv2.drawContours(orig, [c], -1, (0, 0, 255), 2)
            print("lack of enough points to fit ellipse")
    
 
    
    contours_rec_sorted = [x for _, x in sorted(zip(area_rec, contours_rec), key=lambda pair: pair[0])]
    
    #cmap = get_cmap(len(contours_rec_sorted)) 
    
    cmap = get_cmap(len(contours_rec_sorted)+1)
    
    
    tracking_backgd = np.zeros(gray.shape, dtype = "uint8")
    #backgd.fill(128)
    
    label_trait = orig
    track_trait = orig
    #clean area record list
    area_rec = []
    #individual leaf traits sorting based on area order 
    ################################################################################
    for i in range(len(contours_rec_sorted)):
        
        c = contours_rec_sorted[i]
        
        #assign unique color value in opencv format
        color_rgb = tuple(reversed(cmap(i)[:len(cmap(i))-1]))
        
        color_rgb = tuple([255*x for x in color_rgb])
        
        
        # get coordinates of bounding box
        
        #(x,y,w,h) = cv2.boundingRect(c)
        
        rect = cv2.minAreaRect(c)
        box = cv2.boxPoints(rect)
        box = np.array(box, dtype="int")
        box_coord_flat = box.flatten()

        box_coord = []
        for item in box_coord_flat:
            box_coord.append(item)
            
        #box_coord_list = list(map(int,box_coord.split()))
        #print(type(box_coord))
        #print("bbox coordinates :{0}".format(box_coord))
        
        
        
        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        #label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        
        #draw filled contour
        #label_trait = cv2.drawContours(orig, [c], -1, color_rgb, -1)
        
        label_trait = cv2.drawContours(orig, [c], -1, color_rgb, 2)
        
        label_trait = cv2.putText(orig, "#{}".format(i+1), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color_rgb, 1)
        #label_trait = cv2.putText(backgd, "#{}".format(i+1), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color_rgb, 1)
        
        #draw mini bounding box
        #label_trait = cv2.drawContours(orig, [box], -1, (0, 255, 0), 2)
        
        #######################################individual leaf curvature computation
        
        #Get rotated bounding ellipse of contour
        ellipse = cv2.fitEllipse(c)
        
        #get paramters of ellipse
        ((xc,yc), (d1,d2), angle) = ellipse
        
        # draw circle at ellipse center
        #label_trait = cv2.ellipse(orig, ellipse, color_rgb, 2)
        #label_trait = cv2.circle(backgd, (int(xc),int(yc)), 10, color_rgb, -1)
        
        track_trait = cv2.circle(tracking_backgd, (int(xc),int(yc)), 5, (255, 255, 255), -1)
        
        
        #draw major radius
        #compute major radius
        rmajor = max(d1,d2)/2
        rminor = min(d1,d2)/2
        
        if angle > 90:
            angle = angle - 90
        else:
            angle = angle + 90
        
        #print(angle)
        
        xtop = xc + math.cos(math.radians(angle))*rmajor
        ytop = yc + math.sin(math.radians(angle))*rmajor
        xbot = xc + math.cos(math.radians(angle+180))*rmajor
        ybot = yc + math.sin(math.radians(angle+180))*rmajor
        
        label_trait = cv2.line(orig, (int(xtop),int(ytop)), (int(xbot),int(ybot)), color_rgb, 1)
                
        c_np = np.vstack(c).squeeze()
        
        x = c_np[:,0]
        y = c_np[:,1]
        
        comp_curv = ComputeCurvature(x, y)
        
        curvature = comp_curv.fit(x, y)
        
        #compute solidity
        solidity = float(cv2.contourArea(c))/cv2.contourArea(cv2.convexHull(c))
        
        #print("solidity = {0:.2f}... \n".format(solidity))
        
        
        #record all traits 
        leaf_index_rec.append(i)
        area_rec.append(cv2.contourArea(c))
        curv_rec.append(curvature)
        
        solidity_rec.append(solidity)
        major_axis_rec.append(rmajor)
        minor_axis_rec.append(rminor)
        
        box_coord_rec.append(box_coord)
    ################################################################################
    
    
    #print('unique labels={0}, len(contours_rec)={1}, len(leaf_index_rec)={2}'.format(np.unique(labels),len(contours_rec),len(leaf_index_rec)))
        
    n_contours = len(contours_rec_sorted)
    
    if n_contours > 0:
        print('average curvature = {0:.2f}\n'.format(sum(curv_rec)/n_contours))
    else:
        n_contours = 1.0
    
    
    #print(leaf_color_ratio_rec)
    
    return sum(curv_rec)/n_contours, label_trait, track_trait, leaf_index_rec, contours_rec, area_rec, curv_rec, solidity_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec, box_coord_rec
    
    


def RGB2HEX(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))
    

def RGB2FLOAT(color):
    return "{:.2f}{:.2f}{:.2f}".format(int(color[0]/255.0), int(color[1]/255.0), int(color[2]/255.0))



'''
def color_quantization(image, mask, save_path, num_clusters):
    
    #grab image width and height
    (h, w) = image.shape[:2]
    
    #change the color storage order
    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    #apply the mask to get the segmentation of plant
    masked_image = cv2.bitwise_and(image, image, mask = mask)
       
    # reshape the image to be a list of pixels
    pixels = masked_image.reshape((masked_image.shape[0] * masked_image.shape[1], 3))
        
    ############################################################
    #Clustering process
    ###############################################################
    # cluster the pixel intensities
    clt = MiniBatchKMeans(n_clusters = num_clusters)
    #clt = KMeans(n_clusters = args["clusters"])
    clt.fit(pixels)

    #assign labels to each cluster 
    labels = clt.fit_predict(pixels)

    #obtain the quantized clusters using each label
    quant = clt.cluster_centers_.astype("uint8")[labels]

    # reshape the feature vectors to images
    quant = quant.reshape((h, w, 3))
    image_rec = pixels.reshape((h, w, 3))
    
    # convert from L*a*b* to RGB
    quant = cv2.cvtColor(quant, cv2.COLOR_RGB2BGR)
    image_rec = cv2.cvtColor(image_rec, cv2.COLOR_RGB2BGR)
    
    # display the images 
    #cv2.imshow("image", np.hstack([image_rec, quant]))
    #cv2.waitKey(0)
    
    #define result path for labeled images
    result_img_path = save_path + 'cluster_out.png'
    
    # save color_quantization results
    cv2.imwrite(result_img_path,quant)

    #Get colors and analze them from masked image
    counts = Counter(labels)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = clt.cluster_centers_
    
    #print(type(center_colors))

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]
    
    
    #######################################################################################
    threshold = 60
    
    selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))

    for i in range(num_clusters):
        curr_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[i]]]))) 
        diff = deltaE_cie76(selected_color, curr_color)
        if (diff < threshold):
            print("Color difference value is : {0} \n".format(str(diff)))
    ###########################################################################################
    #print(hex_colors)
    
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    
    #print(index_bkg[0])

    #print(counts)
    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)
        
    # build a histogram of clusters and then create a figure representing the number of pixels labeled to each color
    hist = utils.centroid_histogram(clt)

    # remove the background color cluster
    clt.cluster_centers_ = np.delete(clt.cluster_centers_, index_bkg[0], axis=0)
    
    #build a histogram of clusters using center lables
    numLabels = utils.plot_centroid_histogram(save_path,clt)

    #create a figure representing the distribution of each color
    bar = utils.plot_colors(hist, clt.cluster_centers_)

    #save a figure of color bar 
    utils.plot_color_bar(save_path, bar)

    return rgb_colors
'''

def get_cmap(n, name = 'hsv'):
    '''Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
    RGB color; the keyword argument name must be a standard mpl colormap name.'''
    return plt.cm.get_cmap(name, n)
    
    

def color_region(image, mask, save_path, num_clusters):
    
    # read the image
     #grab image width and height
    (h, w) = image.shape[:2]

    #apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask = mask)
    
    #define result path for labeled images
    result_img_path = save_path + 'masked.png'
    cv2.imwrite(result_img_path, masked_image_ori)
    
    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))
    
    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    #num_clusters = 5
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)


    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    #define result path for labeled images
    result_img_path = save_path + 'clustered.png'
    cv2.imwrite(result_img_path, segmented_image_BRG)


    '''
    fig = plt.figure()
    ax = Axes3D(fig)        
    for label, pix in zip(labels, segmented_image):
        ax.scatter(pix[0], pix[1], pix[2], color = (centers))
            
    result_file = (save_path + base_name + 'color_cluster_distributation.png')
    plt.savefig(result_file)
    '''
    #Show only one chosen cluster 
    #masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    #cluster = 2

    cmap = get_cmap(num_clusters+1)
    
    #clrs = sns.color_palette('husl', n_colors = num_clusters)  # a list of RGB tuples

    color_conversion = interp1d([0,1],[0,255])


    for cluster in range(num_clusters):

        print("Processing color cluster {0} ...\n".format(cluster))
        #print(clrs[cluster])
        #print(color_conversion(clrs[cluster]))

        masked_image[labels_flat == cluster] = centers[cluster]

        #print(centers[cluster])

        #convert back to original shape
        masked_image_rp = masked_image.reshape(image_RGB.shape)

        #masked_image_BRG = cv2.cvtColor(masked_image, cv2.COLOR_RGB2BGR)
        #cv2.imwrite('maksed.png', masked_image_BRG)

        gray = cv2.cvtColor(masked_image_rp, cv2.COLOR_BGR2GRAY)

        # threshold the image, then perform a series of erosions +
        # dilations to remove any small regions of noise
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)[1]

        #thresh_cleaned = clear_border(thresh)

        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        #c = max(cnts, key=cv2.contourArea)

        '''
        # compute the center of the contour area and draw a circle representing the center
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        # draw the countour number on the image
        result = cv2.putText(masked_image_rp, "#{}".format(cluster + 1), (cX - 20, cY), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
        '''
        
        
        if not cnts:
            print("findContours is empty")
        else:
            
            # loop over the (unsorted) contours and draw them
            for (i, c) in enumerate(cnts):

                result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(np.random.random(3)), 2)
                #result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(clrs[cluster]), 2)

            #result = result(np.where(result == 0)== 255)
            result[result == 0] = 255


            result_BRG = cv2.cvtColor(result, cv2.COLOR_RGB2BGR)
            result_img_path = save_path + 'result_' + str(cluster) + '.png'
            cv2.imwrite(result_img_path, result_BRG)


    
    counts = Counter(labels_flat)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))
    
    center_colors = centers

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [RGB2FLOAT(ordered_colors[i]) for i in counts.keys()]

    #print(hex_colors)
    
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    
    #print(index_bkg[0])

    #print(counts)
    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)

   
    return rgb_colors, counts, hex_colors


def _normalise_image(image, *, image_cmap=None):
    image = img_as_float(image)
    if image.ndim == 2:
        if image_cmap is None:
            image = gray2rgb(image)
        else:
            image = plt.get_cmap(image_cmap)(image)[..., :3]
    return image


'''
def overlay_skeleton_endpoints(image, stats, *, image_cmap=None, axes=None):

    image = _normalise_image(image, image_cmap=image_cmap)
    summary = stats
    # transforming from row, col to x, y
    #coords_cols = (['image-coord-src-%i' % i for i in [1, 0]] +
    #               ['image-coord-dst-%i' % i for i in [1, 0]])
    
    coords_cols_src = (['image-coord-src-%i' % i for i in [1, 0]])
    coords_cols_dst = (['image-coord-dst-%i' % i for i in [1, 0]])
    
    #coords = summary[coords_cols].values.reshape((-1, 1, 2))
    
    coords_src = summary[coords_cols_src].values
    coords_dst = summary[coords_cols_dst].values

    coords_src_x = [i[0] for i in coords_src]
    coords_src_y = [i[1] for i in coords_src]
    
    coords_dst_x = [i[0] for i in coords_dst]
    coords_dst_y = [i[1] for i in coords_dst]
    
    img_marker = np.zeros_like(image, dtype = np.uint8)
    img_marker.fill(0) # or img[:] = 255
    img_marker[list(map(int, coords_src_y)), list(map(int, coords_src_x))] = 255
    img_marker[list(map(int, coords_dst_y)), list(map(int, coords_dst_x))] = 255
    
    #print("img_marker")
    #print(img_marker.shape)
    
    if axes is None:
        fig, axes = plt.subplots()
    
    axes.axis('off')
    axes.imshow(image)

    axes.scatter(coords_src_x, coords_src_y, c = 'w')
    axes.scatter(coords_dst_x, coords_dst_y, c = 'w')

    return fig, img_marker
    #return coords
'''

def outlier_doubleMAD(data,thresh = 3.5):
    
    """
    FOR ASSYMMETRIC DISTRIBUTION
    Returns : filtered array excluding the outliers

    Parameters : the actual data Points array

    Calculates median to divide data into 2 halves.(skew conditions handled)
    Then those two halves are treated as separate data with calculation same as for symmetric distribution.(first answer) 
    Only difference being , the thresholds are now the median distance of the right and left median with the actual data median
    """
    
    # warning: this function does not check for NAs
    # nor does it address issues when 
    # more than 50% of your data have identical values
    m = np.median(data)
    abs_dev = np.abs(data - m)
    left_mad = np.median(abs_dev[data <= m])
    right_mad = np.median(abs_dev[data >= m])
    data_mad = left_mad * np.ones(len(data))
    data_mad[data > m] = right_mad
    modified_z_score = 0.6745 * abs_dev / data_mad
    modified_z_score[data == m] = 0
    return modified_z_score > thresh



# Convert it to LAB color space to access the luminous channel which is independent of colors.
def isbright(image_file):
    
    # Set up threshold value for luminous channel, can be adjusted and generalized 
    thresh = 1.5
    
    # Load image file 
    orig = cv2.imread(image_file)
    
    # Make backup image
    image = orig.copy()
    
    # Get file name
    #abs_path = os.path.abspath(image_file)
    
    #filename, file_extension = os.path.splitext(abs_path)
    #base_name = os.path.splitext(os.path.basename(filename))[0]

    image_file_name = Path(image_file).name
    
    # Convert color space to LAB format and extract L channel
    L, A, B = cv2.split(cv2.cvtColor(image, cv2.COLOR_BGR2LAB))
    
    # Normalize L channel by dividing all pixel values with maximum pixel value
    L = L/np.max(L)
    
    text_bool = "bright" if np.mean(L) < thresh else "dark"
    
    #return image_file_name, np.mean(L), text_bool
    
    print("np.mean(L) < thresh = {}".format(np.mean(L)))
    
    #return np.mean(L) < thresh
    
    return 1.0 < thresh
    

def remove_character_string(str_input):
    
    return str_input.replace('#', '')



def hex_mean_color(color1, color2, color3, color4):
    
    average_value = (int(remove_character_string(color1), 16) + int(remove_character_string(color2), 16) + int(remove_character_string(color3), 16) + int(remove_character_string(color4), 16))//4
       
    return hex(average_value)


def region_extracted(orig, x, y, w, h):
    
    """compute rect region based on left top corner coordinates and dimension of the region
    
    Inputs: 
    
        orig: image
        
        x, y: left top corner coordinates 
        
        w, h: dimension of the region

    Returns:
    
        roi: region of interest
        
    """   
    roi = orig[y:y+h, x:x+w]
    
    return roi





def extract_traits(image_file):


    #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    base_name = os.path.splitext(os.path.basename(filename))[0]

    file_size = os.path.getsize(image_file)/MBFACTOR
    
    image_file_name = Path(image_file).name
   
    
    # make the folder to store the results
    #current_path = abs_path + '/'
    
    print("Exacting traits for image : {0}\n".format(str(image_file_name)))
     
    # save folder construction
    if (args['result']):
        save_path = args['result']
    else:
         # save folder construction
        #mkpath = os.path.dirname(abs_path) +'/marker_detection'
        #mkdir(mkpath)
        #marker_save_path = mkpath + '/'
        
        mkpath = os.path.dirname(abs_path) + '/' + base_name + '/plant_result'
        mkdir(mkpath)
        save_path = mkpath + '/'
        
        mkpath = os.path.dirname(abs_path) + '/' + base_name + '/color_checker_result'
        mkdir(mkpath)
        color_checker_path = mkpath + '/'
        
       
    print("results_folder: {0}".format(str(save_path)))
    
    print("color_checker_path:  {0}\n".format(str(color_checker_path)))
    
    
    ##############################

    
        
    if isbright(image_file):
    
        if (file_size > 5.0):
            print("File size is {0} MB\n".format(str(int(file_size))))
        else:
            print("Plant object segmentation using automatic color clustering method... \n")
        
        image = cv2.imread(image_file)
        
        # make backup image
        orig = image.copy()
        
        # get the dimension of the image
        img_height, img_width, img_channels = orig.shape
    
        source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
        

         
        args_colorspace = args['color_space']
        args_channels = args['channels']
        args_num_clusters = args['num_clusters']
        
        ##########################################################################
        #barcode detection and recognition
        x = int(0)
        y = int(0)
        w = int(img_width*0.3)
        h = int(img_height*0.3)

        roi_image_barcode = region_extracted(orig, x, y, w, h)
        
        #tag_info = barcode_detect(roi_image)
        
        # initialize the cv2 QRCode detector
        detector = cv2.QRCodeDetector()
        
        # detect and decode
        (QR_data, bbox, straight_qrcode) = detector.detectAndDecode(roi_image_barcode)
        
        print("QR code info: {}\n".format(QR_data))
        
        ####################################################
        #Color checker detection
        
        #define color checker region
        x = int(img_width*0.3)
        y = int(0)
        w = int(img_width*0.7)
        h = int(img_height*0.4)

        roi_image_checker = region_extracted(orig, x, y, w, h)
        
        # save result
        #result_file = (save_path + base_name + '_color_checker' + file_extension)
        #cv2.imwrite(result_file, roi_image_checker)
        
        
        #orig = sticker_crop_img.copy()
        
        #color clustering based plant object segmentation
        thresh_checker = color_cluster_seg(roi_image_checker.copy(), args_colorspace, args_channels, args_num_clusters)
        
        (nb_checker, output, stats, centroids) = cv2.connectedComponentsWithStats(thresh_checker, connectivity = 8)
        
        print("Number of color checkers detected: {} \n".format(nb_checker-1))
        
        # find contours in the thresholded image
        cnts = cv2.findContours(thresh_checker.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        cnts = imutils.grab_contours(cnts)
        
        # initialize the shape detector and color labeler
        cl = ColorLabeler()
        
        lab = cv2.cvtColor(roi_image_checker.copy(), cv2.COLOR_BGR2LAB)
        
        # loop over the contours
        for c in cnts:
            # compute the center of the contour
            M = cv2.moments(c)
            cX = int((M["m10"] / M["m00"]))
            cY = int((M["m01"] / M["m00"]))
            # detect the shape of the contour and label the color

            color = cl.label(lab, c)
            # multiply the contour (x, y)-coordinates by the resize ratio,
            # then draw the contours and the name of the shape and labeled
            # color on the image
            c = c.astype("float")
            c = c.astype("int")
            text = "{}".format(color)
            cv2.drawContours(roi_image_checker, [c], -1, (0, 255, 0), 2)
            cv2.putText(roi_image_checker, text, (cX-80, cY), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
        
        result_file = (color_checker_path + base_name + '_color_checker_detected' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, roi_image_checker)
        
        # apply individual object mask
        color_checker_masked = cv2.bitwise_and(roi_image_checker.copy(), roi_image_checker.copy(), mask = thresh_checker)
        
        #thresh = ~thresh
        # save segmentation result
        result_file = (color_checker_path + base_name + '_color_checker_seg' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, color_checker_masked)

        
        num_clusters = nb_checker + 1
        #save color quantization result
        #rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
        (rgb_colors_checker, counts_checker, hex_colors_checker) = color_region(color_checker_masked.copy(), thresh_checker, color_checker_path, num_clusters)
        
        
         #find external contour 
        (trait_img_checker, area_checker, solidity_checker, max_width_checker, max_height_checker) = comp_external_contour(roi_image_checker.copy(), thresh_checker)
        # save segmentation result
        result_file = (color_checker_path + base_name + '_color_checker__excontour' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, trait_img_checker)   

        #compute the diagonal path length of each color checker
        avg_diagonal_length = np.average([max_width_checker, max_height_checker])
        #avg_diagonal_length = 237.0, max_width_checker = 235, max_height_checker = 239
        print("avg_diagonal_length = {}, max_width_checker = {}, max_height_checker = {}\n".format(avg_diagonal_length, max_width_checker, max_height_checker))
         
        

        ##########################################################################
        

        (sticker_crop_img) = tray_detect(orig, template_tray)
        
        # save segmentation result
        result_file = (save_path + base_name + '_tray_detection.' + args['filetype'])
        #print(result_file)
        cv2.imwrite(result_file, sticker_crop_img)

        '''
        #Plant region detection
        x = int(0)
        y = int(img_height*0.4)
        w = int(img_width*1.0)
        h = int(img_height*0.6)

        roi_image = region_extracted(orig, x, y, w, h)
        '''
        roi_image = sticker_crop_img
        
        # save result
        result_file = (save_path + base_name + '_plant_region' + file_extension)
        cv2.imwrite(result_file, roi_image)
        
        #roi_image = orig.copy()
        
        orig = roi_image.copy()
        
        #orig = sticker_crop_img.copy()
        
        #color clustering based plant object segmentation
        thresh = color_cluster_seg(orig, args_colorspace, args_channels, args_num_clusters)
        
        #thresh = ~thresh
        # save segmentation result
        result_file = (save_path + base_name + '_seg' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, thresh)

        
        num_clusters = 5
        #save color quantization result
        #rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
        (rgb_colors, counts, hex_colors) = color_region(orig, thresh, save_path, num_clusters)
        
        
         #find external contour 
        (trait_img, area, solidity, max_width, max_height) = comp_external_contour(orig, thresh)
        # save segmentation result
        result_file = (save_path + base_name + '_excontour' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, trait_img)   
        

        #print("hex_colors = {} {}\n".format(hex_colors, type(hex_colors)))
        
        list_counts = list(counts.values())
        
        #list_hex_colors = list(hex_colors)
        
        #print(type(list_counts))
        
        color_ratio = []
        
        for value_counts, value_hex in zip(list_counts, hex_colors):
            
            #print(percentage(value, np.sum(list_counts)))
            
            color_ratio.append(percentage(value_counts, np.sum(list_counts)))
            
            #print("value_hex = {0}".format(value_hex))
            
            #value_hex.append(value_hex)
            
            
        #color_ratio_rec.append(color_ratio)
        #color_value_rec.append(hex_colors)
        
        
        
        #selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))
        

        
        ###############################################
        
        #accquire medial axis of segmentation mask
        #image_skeleton = medial_axis_image(thresh)
        
        image_skeleton, skeleton = skeleton_bw(thresh)

        # save _skeleton result
        result_file = (save_path + base_name + '_skeleton' + file_extension)
        cv2.imwrite(result_file, img_as_ubyte(image_skeleton))

        
        ############################################## leaf number computation

        #min_distance_value = 3
            
        print("min_distance_value = {}\n".format(min_distance_value))
        
        #watershed based leaf area segmentaiton 
        labels = watershed_seg(orig, thresh, min_distance_value)
        
        #n_leaves = int(len(np.unique(labels)))
        

        
        #labels = watershed_seg_marker(orig, thresh, min_distance_value, img_marker)
        
        #individual_object_seg(orig, labels, save_path, base_name, file_extension)

        #save watershed result label image
        #Map component labels to hue val
        label_hue = np.uint8(128*labels/np.max(labels))
        #label_hue[labels == largest_label] = np.uint8(15)
        blank_ch = 255*np.ones_like(label_hue)
        labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

        # cvt to BGR for display
        labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

        # set background label to black
        labeled_img[label_hue==0] = 0
        result_file = (save_path + base_name + '_label' + file_extension)
        #plt.imsave(result_file, img_as_float(labels), cmap = "Spectral")
        cv2.imwrite(result_file, labeled_img)
        
        
        #(avg_curv, label_trait, track_trait, leaf_index_rec, contours_rec, area_rec, curv_rec, solidity_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec, box_coord_rec) = leaf_traits_computation(roi_image.copy(), labels, save_path, base_name, file_extension)
        

        #################################################################end of validation file
        n_leaves = int(len(np.unique(labels)))
        
        #n_leaves = int(len((leaf_index_rec)))
        
        print('number of leaves = {0}'.format(n_leaves))
        
        #save watershed result label image
        #result_file = (save_path + base_name + '_leafspec' + file_extension)
        #cv2.imwrite(result_file, label_trait)
        
        #save watershed result label image
        #result_file = (track_save_path + base_name + '_trace' + file_extension)
        #cv2.imwrite(result_file, track_trait)
        
       
      
        
    else:
        
        area=solidity=max_width=max_height=avg_curv=n_leaves=0
        
    #print("[INFO] {} n_leaves found\n".format(len(np.unique(labels)) - 1))
    
    #Path("/tmp/d/a.dat").name
    
    #print("color_ratio = {}".format(color_ratio))
    
    #print("hex_colors = {}".format(hex_colors))
    
    #return image_file_name, QR_data, area, solidity, max_width, max_height, avg_curv, n_leaves, color_ratio, hex_colors, leaf_index_rec, area_rec, curv_rec, solidity_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec
    
    longest_axis = max(max_width, max_height)
    
    #return image_file_name, QR_data, area, solidity, max_width, max_height, n_leaves, color_ratio, hex_colors
    
    return image_file_name, QR_data, area, solidity, longest_axis, avg_diagonal_length, n_leaves, color_ratio, hex_colors





if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help="path to image file")
    ap.add_argument("-ft", "--filetype", required=True,    help="Image filetype")
    ap.add_argument("-r", "--result", required = False,    help="result path")
    ap.add_argument('-s', '--color-space', type = str, required = False, default ='lab', help='Color space to use: BGR, HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type = str, required = False, default='1', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type = int, required = False, default = 2,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-min', '--min_size', type = int, required = False, default = 100,  help = 'min size of object to be segmented.')
    ap.add_argument('-md', '--min_dist', type = int, required = False, default = 50,  help = 'distance threshold of watershed segmentation.')
    ap.add_argument('-tp', '--temp_path', required = False,  default ='/marker_template/tray.tif',  help = "template for tray")
    args = vars(ap.parse_args())
    
    
    # setting path to model file
    file_path = args["path"]
    ext = args['filetype']
    
    min_size = args['min_size']
    min_distance_value = args['min_dist']

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    global  template
    template_path = file_path + args['temp_path']

    
    #setup marker path to load template
    template_path = file_path + args['temp_path']


    try:
        # check to see if file is readable
        with open(template_path) as tempFile:

            # Read the template 
            template_tray = cv2.imread(template_path, 0)
            print("Tray template loaded successfully...\n")
            template_option = 1
            
            
    except IOError as err:
        
        print("Error reading Barcode template file {0}: {1}\n".format(barcode_path, err))
        
        #exit(0)

    #print((imgList))
    #global save_path
    
    n_images = len(imgList)
    
    result_list = []
    
    result_list_leaf = []
    
    #loop execute
    for image_id, image in enumerate(imgList):
        
    
        #(filename, QR_data, area, solidity, max_width, max_height, avg_curv, n_leaves, color_ratio, hex_colors, leaf_index_rec, area_rec, curv_rec, solidity_rec, major_axis_rec, minor_axis_rec, leaf_color_ratio_rec, leaf_color_value_rec) = extract_traits(image)
        
        
        (filename, QR_data, area, solidity, longest_axis, avg_diagonal_length, n_leaves, color_ratio, hex_colors) = extract_traits(image)
        
        #result_list.append([filename, area, solidity, max_width, max_height, avg_curv, n_leaves, color_ratio[0], color_ratio[1], color_ratio[2], color_ratio[3], hex_colors[0], hex_colors[1], hex_colors[2], hex_colors[3]])
        
        result_list.append([filename, QR_data, area, solidity, longest_axis, avg_diagonal_length, n_leaves, color_ratio[0], color_ratio[1], color_ratio[2], color_ratio[3], 
                            str(matplotlib.colors.to_rgb(hex_colors[0])), str(matplotlib.colors.to_rgb(hex_colors[1])), str(matplotlib.colors.to_rgb(hex_colors[2])), str(matplotlib.colors.to_rgb(hex_colors[3])),
                            hex_colors[0], hex_colors[1], hex_colors[2], hex_colors[3], hex_mean_color(hex_colors[0], hex_colors[1], hex_colors[2], hex_colors[3])])
        '''
        for i in range(len(leaf_index_rec)):
                        #result_list_leaf.append([filename, leaf_index_rec[i], area_rec[i], curv_rec[i], solidity_rec[i], major_axis_rec[i], minor_axis_rec[i], leaf_color_ratio_rec[i][0], leaf_color_ratio_rec[i][1], leaf_color_ratio_rec[i][2], leaf_color_ratio_rec[i][3], leaf_color_value_rec[i][0],leaf_color_value_rec[i][1],leaf_color_value_rec[i][2],leaf_color_value_rec[i][3]])
            
            result_list_leaf.append([filename, QR_data, leaf_index_rec[i], area_rec[i], curv_rec[i], solidity_rec[i], major_axis_rec[i], minor_axis_rec[i], leaf_color_ratio_rec[i][0], leaf_color_ratio_rec[i][1], leaf_color_ratio_rec[i][2], leaf_color_ratio_rec[i][3], 
                                    str(matplotlib.colors.to_rgb(leaf_color_value_rec[i][0])), str(matplotlib.colors.to_rgb(leaf_color_value_rec[i][1])), str(matplotlib.colors.to_rgb(leaf_color_value_rec[i][2])), str(matplotlib.colors.to_rgb(leaf_color_value_rec[i][3])),
                                    str(leaf_color_value_rec[i][0]), str(leaf_color_value_rec[i][1]), str(leaf_color_value_rec[i][2]), str(leaf_color_value_rec[i][3]),
                                    hex_mean_color(leaf_color_value_rec[i][0], leaf_color_value_rec[i][1], leaf_color_value_rec[i][2], leaf_color_value_rec[i][3])])
        '''
    ########################################################################
    '''
    # get cpu number for parallel processing
    agents = psutil.cpu_count()   
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result_list = pool.map(extract_traits, imgList)
        pool.terminate()
    '''
    
    #########################################################################
    #trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
 
    #table = tabulate(result_list, headers = ['filename', 'area', 'solidity', 'max_width', 'max_height' ,'avg_curv', 'n_leaves', 'cluster 1', 'cluster 2', 'cluster 3', 'cluster 4', 'cluster 1 hex value', 'cluster 2 hex value', 'cluster 3 hex value', 'cluster 4 hex value'], tablefmt = 'orgtbl')

    #print(table + "\n")
    
    
    '''
    print("Summary: Leaf specific traits...\n")
    
    #output in command window in a sum table
 
    table = tabulate(result_list_leaf, headers = ['filename', 'leaf_index', 'area', 'curvature', 'solidity', 'major_axis', 'minor_axis', 'cluster 1', 'cluster 2', 'cluster 3', 'cluster 4', 'cluster 1 hex value', 'cluster 2 hex value', 'cluster 3 hex value', 'cluster 4 hex value'], tablefmt = 'orgtbl')

    print(table + "\n")
    '''
    
    
    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')
        trait_file_csv = (args['result'] + 'trait.csv')
    else:
        trait_file = (file_path + 'trait.xlsx')
        trait_file_csv = (file_path + 'trait.csv')
    
    
    if os.path.isfile(trait_file):
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet = wb.active
        
        sheet.delete_rows(2, sheet.max_row+1) # for entire sheet
        
        #sheet_leaf = wb.create_sheet()
        
        #sheet_leaf.delete_rows(2, sheet_leaf.max_row+1) # for entire sheet
        
        

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        sheet_leaf = wb.create_sheet()

        sheet.cell(row = 1, column = 1).value = 'filename'
        sheet.cell(row = 1, column = 2).value = 'QR_data'
        sheet.cell(row = 1, column = 3).value = 'leaf_area'
        sheet.cell(row = 1, column = 4).value = 'solidity'
        sheet.cell(row = 1, column = 5).value = 'longest_axis'
        sheet.cell(row = 1, column = 6).value = 'avg_diagonal_length'
        sheet.cell(row = 1, column = 7).value = 'number_leaf'
        sheet.cell(row = 1, column = 8).value = 'color distribution cluster 1'
        sheet.cell(row = 1, column = 9).value = 'color distribution cluster 2'
        sheet.cell(row = 1, column = 10).value = 'color distribution cluster 3'
        sheet.cell(row = 1, column = 11).value = 'color distribution cluster 4'
        sheet.cell(row = 1, column = 12).value = 'color cluster 1 RGB value'
        sheet.cell(row = 1, column = 13).value = 'color cluster 2 RGB value'
        sheet.cell(row = 1, column = 14).value = 'color cluster 3 RGB value'
        sheet.cell(row = 1, column = 15).value = 'color cluster 4 RGB value'
        sheet.cell(row = 1, column = 16).value = 'color cluster 1 HEX value'
        sheet.cell(row = 1, column = 17).value = 'color cluster 2 HEX value'
        sheet.cell(row = 1, column = 18).value = 'color cluster 3 HEX value'
        sheet.cell(row = 1, column = 19).value = 'color cluster 4 HEX value'
        sheet.cell(row = 1, column = 20).value = 'average HEX value'       
        
    
        '''
        sheet_leaf.cell(row = 1, column = 1).value = 'filename'
        sheet_leaf.cell(row = 1, column = 1).value = 'filename'
        sheet_leaf.cell(row = 1, column = 2).value = 'leaf_index'
        sheet_leaf.cell(row = 1, column = 3).value = 'area'
        sheet_leaf.cell(row = 1, column = 4).value = 'curvature'
        sheet_leaf.cell(row = 1, column = 5).value = 'solidity'
        sheet_leaf.cell(row = 1, column = 6).value = 'major_axis'
        sheet_leaf.cell(row = 1, column = 7).value = 'minor_axis'
        sheet_leaf.cell(row = 1, column = 8).value = 'color distribution cluster 1'
        sheet_leaf.cell(row = 1, column = 9).value = 'color distribution cluster 2'
        sheet_leaf.cell(row = 1, column = 10).value = 'color distribution cluster 3'
        sheet_leaf.cell(row = 1, column = 11).value = 'color distribution cluster 4'
        sheet_leaf.cell(row = 1, column = 12).value = 'color cluster 1 RGB value'
        sheet_leaf.cell(row = 1, column = 13).value = 'color cluster 2 RGB value'
        sheet_leaf.cell(row = 1, column = 14).value = 'color cluster 3 RGB value'
        sheet_leaf.cell(row = 1, column = 15).value = 'color cluster 4 RGB value'
        sheet_leaf.cell(row = 1, column = 16).value = 'color cluster 1 HEX value'
        sheet_leaf.cell(row = 1, column = 17).value = 'color cluster 2 HEX value'
        sheet_leaf.cell(row = 1, column = 18).value = 'color cluster 3 HEX value'
        sheet_leaf.cell(row = 1, column = 19).value = 'color cluster 4 HEX value'
        sheet_leaf.cell(row = 1, column = 20).value = 'average HEX value'      
        
        '''
        
    for row in result_list:
        sheet.append(row)
   
    #for row in result_list_leaf:
        #sheet_leaf.append(row)
    
    '''
    for row in result_list_leaf:
        sheet_leaf.append(row)
    '''
    #save the csv file
    wb.save(trait_file)
    
    
    
    '''
    wb = openpyxl.load_workbook(trait_file)
    sh = wb.active # was .get_active_sheet()
    with open(trait_file_csv, 'w', newline = "") as f:
        c = csv.writer(f)
        for r in sh.rows: # generator; was sh.rows
            c.writerow([cell.value for cell in r])
    '''

    

    
