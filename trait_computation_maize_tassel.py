'''
Name: trait_extract_parallel.py

Version: 1.0

Summary: Extract maize_tassel traits 
    
Author: suxing liu

Author-email: suxingliu@gmail.com

Created: 2022-09-29

USAGE:

time python3 trait_computation_maize_tassel.py -p ~/example/plant_test/seeds/test_tassel/ -ft png 

time python3 trait_computation_maize_tassel.py -p ~/example/plant_test/seeds/test_tassel/ -s HSV -c 1 -ft png

'''

# import necessary packages
import os
import glob
import utils

from collections import Counter

import argparse

from sklearn.cluster import KMeans
from sklearn.cluster import MiniBatchKMeans

from skimage.feature import peak_local_max
from skimage.morphology import medial_axis
from skimage import img_as_float, img_as_ubyte, img_as_bool, img_as_int
from skimage import measure
from skimage.color import rgb2lab, deltaE_cie76
from skimage import morphology
from skimage.segmentation import clear_border, watershed
from skimage.measure import regionprops

from scipy.spatial import distance as dist
from scipy import optimize
from scipy import ndimage
from scipy.interpolate import interp1d

from skan import skeleton_to_csgraph, Skeleton, summarize, draw

import imutils
from imutils import perspective

import numpy as np
import argparse
import cv2

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import collections

import math
import openpyxl
import csv
    
from tabulate import tabulate

from pathlib import Path 
from pylibdmtx.pylibdmtx import decode
import re


import psutil
import concurrent.futures
import multiprocessing
from multiprocessing import Pool
from contextlib import closing

import natsort 

import warnings
warnings.filterwarnings("ignore")



MBFACTOR = float(1<<20)


def mkdir(path):

    """create folder and path to store the output results
    
    Inputs: 
    
        path: result path
        
       
    Returns:
    
        create path and folder if not exist  
        
    """   

    # remove space at the beginning
    path=path.strip()
    # remove slash at the end
    path=path.rstrip("\\")
 
    # path exist?   # True  # False
    isExists=os.path.exists(path)
 
    # process
    if not isExists:
        # construct the path and folder
        #print path + ' folder constructed!'
        # make dir
        os.makedirs(path)
        return True
    else:
        # if exists, return 
        #print path+' path exists!'
        return False
        


def color_cluster_seg(image, args_colorspace, args_channels, args_num_clusters):


    """color clustering based object segmentation
    
    Inputs: 
    
        image: image contains the plant objects
        
        args_colorspace: user-defined color space for clustering 
        
        args_channels: user-defined color channel for clustering 
        
        args_num_clusters: number of clustering
        
    Returns:
    
        img_thresh: mask image with the segmentation of the plant object 
        
    """   
    
    # Change image color space, if necessary.
    colorSpace = args_colorspace.lower()

    if colorSpace == 'hsv':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        
    elif colorSpace == 'ycrcb' or colorSpace == 'ycc':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2YCrCb)
        
    elif colorSpace == 'lab':
        image = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        
    else:
        colorSpace = 'bgr'  # set for file naming purposes
        
        
    #image = cv2.pyrMeanShiftFiltering(image, 21, 70)

    # Keep only the selected channels for K-means clustering.
    if args_channels != 'all':
        channels = cv2.split(image)
        channelIndices = []
        for char in args_channels:
            channelIndices.append(int(char))
        image = image[:,:,channelIndices]
        if len(image.shape) == 2:
            image.reshape(image.shape[0], image.shape[1], 1)
    
    # get the dimension of image        
    (width, height, n_channel) = image.shape
    
    # Flatten the 2D image array into an MxN feature vector, where M is the number of pixels and N is the dimension (number of channels).
    reshaped = image.reshape(image.shape[0] * image.shape[1], image.shape[2])
    
    # Perform K-means clustering.
    if args_num_clusters < 2:
        print('Warning: num-clusters < 2 invalid. Using num-clusters = 2')
    
    #define number of cluster
    numClusters = max(2, args_num_clusters)
    
    # clustering method
    kmeans = KMeans(n_clusters = numClusters, n_init = 40, max_iter = 500).fit(reshaped)
    
    # get lables 
    pred_label = kmeans.labels_
    
    # Reshape result back into a 2D array, where each element represents the corresponding pixel's cluster index (0 to K - 1).
    clustering = np.reshape(np.array(pred_label, dtype=np.uint8), (image.shape[0], image.shape[1]))

    # Sort the cluster labels in order of the frequency with which they occur.
    sortedLabels = sorted([n for n in range(numClusters)],key = lambda x: -np.sum(clustering == x))

    # Initialize K-means grayscale image; set pixel colors based on clustering.
    kmeansImage = np.zeros(image.shape[:2], dtype=np.uint8)
    for i, label in enumerate(sortedLabels):
        kmeansImage[clustering == label] = int(255 / (numClusters - 1)) * i
    
    ret, thresh = cv2.threshold(kmeansImage,0,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    
    #thresh_cleaned = (thresh)
    
    # clean the border of mask image
    if np.count_nonzero(thresh) > 0:
        
        thresh_cleaned = clear_border(thresh)
    else:
        thresh_cleaned = thresh
    
    # get the connected Components in the mask image
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(thresh_cleaned, connectivity = 8)

    
    # stats[0], centroids[0] are for the background label. ignore
    # cv2.CC_STAT_LEFT, cv2.CC_STAT_TOP, cv2.CC_STAT_WIDTH, cv2.CC_STAT_HEIGHT
    # get all connected Components's area value
    sizes = stats[1:, cv2.CC_STAT_AREA]

    # remove background component
    nb_components = nb_components - 1

    #max_size = width*height*0.1
    
    # create an empty mask image and fill the detected connected components
    img_thresh = np.zeros([width, height], dtype=np.uint8)
    
    #for every component in the image, keep it only if it's above min_size
    for i in range(0, nb_components):
        
        if (sizes[i] >= min_size):
        
            img_thresh[output == i + 1] = 255
        

    
    #if mask contains mutiple non-conected parts, combine them into one. 
    contours, hier = cv2.findContours(img_thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) > 1:
        
        print("mask contains mutiple non-conected parts, combine them into one\n")
        
        # create an size 10 kernel
        kernel = np.ones((10,10), np.uint8)

        # image dilation
        dilation = cv2.dilate(img_thresh.copy(), kernel, iterations = 1)
        
        # image closing
        closing = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
        
        # use the final closing result as mask
        img_thresh = closing

    
    # return segmentation mask
    return img_thresh
    
    


def circle_detection(image):

    """Detecting Circles in Images using OpenCV and Hough Circles
    
    Inputs: 
    
        image: image loaded 

    Returns:
    
        circles: detcted circles
        
        circle_detection_img: circle overlayed with image
        
        diameter_circle: diameter of detected circle
        
    """
    
    # create background image for drawing the detected circle
    output = image.copy()
    
    # obtain image dimension
    img_height, img_width, n_channels = image.shape
    
    #backup input image
    circle_detection_img = image.copy()
    
    # change image from RGB to Gray scale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # apply blur filter 
    blurred = cv2.medianBlur(gray, 25)
    
    # setup parameters for circle detection
    
    # This parameter is the inverse ratio of the accumulator resolution to the image resolution 
    #(see Yuen et al. for more details). Essentially, the larger the dp gets, the smaller the accumulator array gets.
    dp = 1.5
    
    #Minimum distance between the center (x, y) coordinates of detected circles. 
    #If the minDist is too small, multiple circles in the same neighborhood as the original may be (falsely) detected. 
    #If the minDist is too large, then some circles may not be detected at all.
    minDist = 100
    
    #Gradient value used to handle edge detection in the Yuen et al. method.
    #param1 = 30
    
    #accumulator threshold value for the cv2.HOUGH_GRADIENT method. 
    #The smaller the threshold is, the more circles will be detected (including false circles). 
    #The larger the threshold is, the more circles will potentially be returned. 
    #param2 = 30  
    
    #Minimum/Maximum size of the radius (in pixels).
    #minRadius = 80
    #maxRadius = 120 
    
    # detect circles in the image
    #circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, 1.2, minDist, param1=param1, param2=param2, minRadius=minRadius, maxRadius=maxRadius)
    
    # detect circles in the image
    circles = cv2.HoughCircles(blurred, cv2.HOUGH_GRADIENT, dp, minDist)
    
    # initialize diameter of detected circle
    diameter_circle = 0
    
    
    circle_center_coord = []
    circle_center_radius = []
    idx_closest = 0
    
    
    # convert the (x, y) coordinates and radius of the circles to integers
    circles = np.round(circles[0, :]).astype("int")
    
    # loop over the (x, y) coordinates and radius of the circles
    for (x, y, r) in circles:
        
        coord = (x, y)
        
        circle_center_coord.append(coord)
        circle_center_radius.append(r)
    
    
    # choose the left bottom circle if more than one circles are detected 
    if len(circles) > 1:
        
        #finding closest point among the center list of the circles to the right-bottom of the image
        idx_closest = closest_center((0 + img_width, 0 + img_height), circle_center_coord)
    
    else:
        
        # ensure at least some circles were found
        if circles is not None and len(circles) > 0:
            idx_closest = 0
    
    #print("idx_closest = {}\n".format(idx_closest))
    
    # draw the circle in the output image, then draw a center
    circle_detection_img = cv2.circle(output, circle_center_coord[idx_closest], circle_center_radius[idx_closest], (0, 255, 0), 4)
    circle_detection_img = cv2.circle(output, circle_center_coord[idx_closest], 5, (0, 128, 255), -1)
    
    # create an empty mask image and fill the detected connected components
    mask = np.zeros([img_width, img_height], dtype = np.uint8)
    
    # compute the diameter of coin
    diameter_circle = circle_center_radius[idx_closest]*2
    
    
    return circles, circle_detection_img, diameter_circle



def percentage(part, whole):
  
    """compute percentage value
    
    Inputs: 
    
        part, whole: the part and whole value
        
       
    Returns:
    
        string type of the percentage in decimals 
        
    """   
  
    percentage = "{:.2f}".format(float(part)/float(whole))

    return str(percentage)



def region_extracted(orig, x, y, w, h):
    
    """compute rect region based on left top corner coordinates and dimension of the region
    
    Inputs: 
    
        orig: image
        
        x, y: left top corner coordinates 
        
        w, h: dimension of the region

    Returns:
    
        roi: region of interest
        
    """   
    roi = orig[y:y+h, x:x+w]
    
    return roi




def midpoint(ptA, ptB):
    
    """compute middle point of two points in 2D coordinates
    
    Inputs: 
    
        ptA, ptB: coordinates of two points
        
    Returns:
    
        coordinates of the middle point
        
    """   
    return ((ptA[0] + ptB[0]) * 0.5, (ptA[1] + ptB[1]) * 0.5)


# compute the parameters of the external contour of the object 
def comp_external_contour(orig, thresh):


    """compute the parameters of the external contour of the plant object 
    
    Inputs: 
    
        orig: image contains the plant objects
        
        thresh: mask of the plant object
        
    Returns:
    
        trait_img: input image overlayed with external contour and bouding box
        
        tassel_area: area occupied by the tassel in the image
        
        cnt_width, cnt_height: width and height of the tassel
        
        tassel_area_ratio: the ratio of area occupied by the tassel and its convex hull area
        
        
    """   
    
    # get the dimension and color channel of the input image    
    img_height, img_width, img_channels = orig.shape
   
    # initialize parameters
    trait_img = orig.copy()
    area = 0
    w=h=0
    

    ####################################################################################
    #find contours and get the external one
    #contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
   
    #RETR_CCOMP
    contours, hierarchy = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)

    # sort the contours based on area from largest to smallest
    contours_sorted = sorted(contours, key = cv2.contourArea, reverse = True)
    
    #print("length of contours = {}\n".format(len(contours)))
    
    # get the max contour in size
    c_max = max(contours, key = cv2.contourArea)
    
    # compute the area of the max contour
    area_c_max = cv2.contourArea(c_max)
    
    # compute the bounding box of the max contour
    (x, y, w, h) = cv2.boundingRect(c_max)
    
    # compute the convex hull of the max contour
    hull = cv2.convexHull(c_max)
    
    # compute the area of convex hull 
    convex_hull_area = cv2.contourArea(hull)
    

    # compute the width and height of the bounding box of the max contour
    cnt_width = w
    
    cnt_height = h
    
    # initialize parameter 
    area_c = 0
    
    area_child_contour_sum = []
    
    
    ###########################################################################
    # compute all the contours/holes inside the max contour and their areas 
    
    for index, c in enumerate(contours_sorted):
       
        # visualize only the external contour and its bounding box
        if index < 1:
            
            # draw a green contour
            trait_img = cv2.drawContours(orig, [c], -1, (0, 255, 0), 2)
            
            # draw a rectangle to visualize the bounding rect
            trait_img = cv2.rectangle(orig, (x, y), (x+w, y+h), (255, 255, 0), 4)

            # compute the center of the contour
            M = cv2.moments(c)
            
            cX = int(M["m10"] / M["m00"])
            cY = int(M["m01"] / M["m00"])
            
            # draw the center of the shape on the image
            #trait_img = cv2.circle(orig, (cX, cY), 7, (255, 255, 255), -1)
            #trait_img = cv2.putText(orig, "center", (cX - 20, cY - 20),cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
            
 
            ###############################################################################

            # compute the four coordinates to get the center of bounding box
            tl = (x, y+h*0.5)
            tr = (x+w, y+h*0.5)
            br = (x+w*0.5, y)
            bl = (x+w*0.5, y+h)
            
            # compute the midpoint between bottom-left and bottom-right coordinates
            (tltrX, tltrY) = midpoint(tl, tr)
            (blbrX, blbrY) = midpoint(bl, br)
            
            # draw the midpoints on the image
            trait_img = cv2.circle(orig, (int(tltrX), int(tltrY)), 15, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, (int(blbrX), int(blbrY)), 15, (255, 0, 0), -1)

            # draw lines between the midpoints
            trait_img = cv2.line(orig, (int(x), int(y+h*0.5)), (int(x+w), int(y+h*0.5)), (255, 0, 255), 6)
            trait_img = cv2.line(orig, (int(x+w*0.5), int(y)), (int(x+w*0.5), int(y+h)), (255, 0, 255), 6)
            
            # compute the convex hull of the contour
            hull = cv2.convexHull(c)
            
            # draw convexhull in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 2)
            
            
        else:
            # visualize all the contour
            trait_img = cv2.drawContours(orig, [c], -1, (0, 255, 0), 2)
            
            # compute the area of each contour
            area_c = cv2.contourArea(c)
            
            # get the list of the all the contour area
            area_child_contour_sum.append(area_c)
            
    
    #print(area_c_max, sum(area_child_contour_sum))
    
    # compute the tassel_area 
    tassel_area = area_c_max - sum(area_child_contour_sum)
    
    #compute the area ratio of tassel area against its convex hull area
    tassel_area_ratio = tassel_area/convex_hull_area
    
    print("Tassel shape info: Width = {0}, Height= {1}, Area = {2}\n".format(w, h, tassel_area))
   
    # return all the traits
    return trait_img, tassel_area, cnt_width, cnt_height, tassel_area_ratio
    



def RGB2HEX(color):
    
    """convert RGB value to HEX format
    
    Inputs: 
    
        color: color in rgb format
        
    Returns:
    
        color in hex format
        
    """   
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))




def get_cmap(n, name = 'hsv'):
    
    """get n kinds of colors from a color palette 
    
    Inputs: 
    
        n: number of colors
        
        name: the color palette choosed
        
    Returns:
    
        plt.cm.get_cmap(name, n): Returns a function that maps each index in 0, 1, ..., n-1 to a distinct 
        RGB color; the keyword argument name must be a standard mpl colormap name. 
        
    """   

    return plt.cm.get_cmap(name, n)
    
    
 
def color_region(image, mask, save_path, num_clusters):

    """dominant color clustering method to compute the color distribution 
    
    Inputs: 
    
        image: image contains different colors
        
        mask: mask of the plant object
        
        save_path: result path
        
        num_clusters: number of clusters for computation 

    Returns:
    
        rgb_colors: center color values in rgb format for every cluster
        
        counts: percentage of each color cluster
        
        hex_colors: center color values in hex format for every cluster
        
        
    """   

    # read the image
     #grab image width and height
    (h, w) = image.shape[:2]

    #apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask = mask)
    
    #define result path for labeled images
    result_img_path = save_path + 'masked.png'
    cv2.imwrite(result_img_path, masked_image_ori)
    
    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))
    
    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    #num_clusters = 5
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)

    # convert image format from RGB to BGR for OpenCV
    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    
    # define result path for labeled images
    result_img_path = save_path + 'clustered.png'
    cv2.imwrite(result_img_path, segmented_image_BRG)

    # Show only one chosen cluster 
    #masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    #cluster = 2
    
    # get the color template
    cmap = get_cmap(num_clusters+1)
    
    #clrs = sns.color_palette('husl', n_colors = num_clusters)  # a list of RGB tuples

    # convert colors format 
    color_conversion = interp1d([0,1],[0,255])

    # loop over all the clusters
    for cluster in range(num_clusters):

        print("Processing color cluster {0} ...\n".format(cluster))
        #print(clrs[cluster])
        #print(color_conversion(clrs[cluster]))
        
        # choose current label image of same cluster
        masked_image[labels_flat == cluster] = centers[cluster]

        #print(centers[cluster])

        #convert back to original shape
        masked_image_rp = masked_image.reshape(image_RGB.shape)

        #masked_image_BRG = cv2.cvtColor(masked_image, cv2.COLOR_RGB2BGR)
        #cv2.imwrite('maksed.png', masked_image_BRG)
        
        # convert the maksed image from BGR to GRAY
        gray = cv2.cvtColor(masked_image_rp, cv2.COLOR_BGR2GRAY)

        # threshold the image
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)[1]

        #thresh_cleaned = clear_border(thresh)

        # get the external contours
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        #c = max(cnts, key=cv2.contourArea)

        # if no contour was found
        if not cnts:
            print("findContours is empty")
        else:
            
            # loop over the (unsorted) contours and draw them
            for (i, c) in enumerate(cnts):
                
                # draw contours on the masked_image_rp
                result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(np.random.random(3)), 2)
                #result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(clrs[cluster]), 2)

            #result = result(np.where(result == 0)== 255)
            result[result == 0] = 255

            # convert the result image from RGB to BGR format for OpenCV
            result_BRG = cv2.cvtColor(result, cv2.COLOR_RGB2BGR)
            result_img_path = save_path + 'result_' + str(cluster) + '.png'
            
            # save result
            cv2.imwrite(result_img_path, result_BRG)


    # sort to ensure correct color percentage
    counts = Counter(labels_flat)
    counts = dict(sorted(counts.items()))
    
    # get all cluster center colors
    center_colors = centers

    # get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [RGB2HEX(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]

    # find the background index
    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']
    

    #remove background color 
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]
    
    # Using dictionary comprehension to find list keys having value . 
    delete = [key for key in counts if key == index_bkg[0]] 
  
    # delete the key 
    for key in delete: del counts[key] 
   
    # save the color distribution pie chart
    fig = plt.figure(figsize = (6, 6))
    plt.pie(counts.values(), labels = hex_colors, colors = hex_colors)

    #define result path for labeled images
    result_img_path = save_path + 'pie_color.png'
    plt.savefig(result_img_path)

   
    return rgb_colors, counts, hex_colors



def barcode_detect(img_ori):

    """Read barcode in the image and decode barcode info
    
    Inputs: 
    
        img_ori: image contains the barcode region
        
    Returns:
    
        tag_info: decoded barcode information  
        
    """   
    
    # get the dimension of the image
    height, width = img_ori.shape[:2]
    
    # decode the barcode info 
    barcode_info = decode((img_ori.tobytes(), width, height))
    
    # if barcode info was not empty
    if len(barcode_info) > 0:
        
        # get the decoded barcode info data value as string
        barcode_str = str(barcode_info[0].data)
        
        #print('Decoded data:', barcode_str)
        #print(decoded_object.rect.top, decoded_object.rect.left)
        #print(decoded_object.rect.width, decoded_object.rect.height)
 
        # accquire the barcode info and remove extra characters
        tag_info = re.findall(r"'(.*?)'", barcode_str, re.DOTALL)
        tag_info = " ".join(str(x) for x in tag_info)
        tag_info = tag_info.replace("'", "")
        
        print("Tag info: {}\n".format(tag_info))
    
    else:
        # print warning if barcode info was empty
        print("barcode information was not readable!\n")
        tag_info = 'Unreadable'
        
    return tag_info
    
    


def marker_detect(img_ori, template, method, selection_threshold):

    """Detect marker in the image
    
    Inputs: 
    
        img_ori: image contains the marker region
        
        template: preload marker template image
        
        method: method used to compute template matching
        
        selection_threshold: thresh value for accept the template matching result

    Returns:
    
        marker_img: matching region image with marker object  
        
        thresh: mask image of the marker region
        
        coins_width_contour: computed width result based on contour of the object 
        
        coins_width_circle: computed width result based on min circle of the object 
        
    """   
    
    # load the image, clone it for output
    img_rgb = img_ori.copy()
      
    # convert it to grayscale 
    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY) 
      
    # store width and height of template in w and h 
    w, h = template.shape[::-1] 
      
    # Perform match operations. 
    #res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF_NORMED)
    #res = cv2.matchTemplate(img_gray, template, cv2.TM_SQDIFF)
    
    # Perform template matching operations. 
    res = cv2.matchTemplate(img_gray, template, cv2.TM_CCOEFF)
    
    
    # Specify a threshold for template detection as selection_threshold
      
    # Store the coordinates of matched area in a numpy array 
    loc = np.where( res >= selection_threshold)   
    
    if len(loc):
    
        # unwarp the template mathcing result
        (y,x) = np.unravel_index(res.argmax(), res.shape)
        
        # get the template matching region coordinates
        (min_val, max_val, min_loc, max_loc) = cv2.minMaxLoc(res)

        (startX, startY) = max_loc
        endX = startX + template.shape[1]
        endY = startY + template.shape[0]
        
        # get the sub image with matching region
        marker_img = img_ori[startY:endY, startX:endX]
        
        marker_overlay = marker_img
        
        
        
        # Draw a rectangle around the matched region. 
        #for pt in zip(*loc[::-1]): 
            
            #marker_overlay = cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,255,0), 1)

    
        # load the marker image, convert it to grayscale
        marker_img_gray = cv2.cvtColor(marker_img, cv2.COLOR_BGR2GRAY) 
    
       
        # load the image and perform pyramid mean shift filtering to aid the thresholding step
        shifted = cv2.pyrMeanShiftFiltering(marker_img, 21, 51)
        
        # convert the mean shift image to grayscale, then apply Otsu's thresholding
        gray = cv2.cvtColor(shifted, cv2.COLOR_BGR2GRAY)
        
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        
        
        # detect contours in the mask and grab the largest one
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        cnts = imutils.grab_contours(cnts)
        
        # get the contour with largest area
        largest_cnt = max(cnts, key=cv2.contourArea)
        
        #print("[INFO] {} unique contours found in marker_img\n".format(len(cnts)))
        
        # compute the radius of the detected coin
        # calculate the center of the contour
        M = cv2.moments(largest_cnt )
        x = int(M["m10"] / M["m00"])
        y = int(M["m01"] / M["m00"])

        # calculate the radius of the contour from area (I suppose it's a circle)
        area = cv2.contourArea(largest_cnt)
        radius = np.sqrt(area/math.pi)
        coins_width_contour = 2* radius
    
        # draw a circle enclosing the object
        ((x, y), radius) = cv2.minEnclosingCircle(largest_cnt) 
        coins_width_circle = 2* radius
        

    return  marker_img, thresh, coins_width_contour, coins_width_circle



def skeleton_bw(thresh):

    """compute skeleton from binary mask image
    
    Inputs: 
    
        image_skeleton: binary mask image

    Returns:
    
        skeleton_img: skeleton image for output
        
        skeleton: skeleton data 
        
    """
    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it
    
    #convert an image from OpenCV to skimage
    #thresh_sk = img_as_float(thresh)

    #image_bw = img_as_bool((thresh_sk))

    #skeleton = morphology.skeletonize(image_bw)
    ##################################
    #define kernel size
    kernel_size = 15
    
    # taking a matrix of size 25 as the kernel
    kernel = np.ones((kernel_size, kernel_size), np.uint8)
    
    # image dilation
    dilation = cv2.dilate(thresh.copy(), kernel, iterations = 1)
        
    # image closing
    image_bw = cv2.morphologyEx(dilation, cv2.MORPH_CLOSE, kernel)
    
    # get the skeleton 
    skeleton = morphology.thin(image_bw)

    # convert the skeleton image to byte 
    skeleton_img = skeleton.astype(np.uint8) * 255

    return skeleton_img, skeleton



def outlier_doubleMAD(data, thresh = 3.5):
    
    """outlier removal 
    Calculates median to divide data into 2 halves.(skew conditions handled)
    Then those two halves are treated as separate data with calculation same as for symmetric distribution.(first answer) 
    Only difference being , the thresholds are now the median distance of the right and left median with the actual data median
    convert the skeleton to graph and analyze the graph
    
    FOR ASSYMMETRIC DISTRIBUTION
    
    Inputs: the actual data Points array and thresh value

    Returns : filtered array excluding the outliers

    """
    
    # warning: this function does not check for NAs
    # nor does it address issues when 
    # more than 50% of your data have identical values
    m = np.median(data)
    abs_dev = np.abs(data - m)
    left_mad = np.median(abs_dev[data <= m])
    right_mad = np.median(abs_dev[data >= m])
    data_mad = left_mad * np.ones(len(data))
    data_mad[data > m] = right_mad
    modified_z_score = 0.6745 * abs_dev / data_mad
    modified_z_score[data == m] = 0
    return modified_z_score > thresh



def skeleton_graph(image_skeleton):
    
    """convert the skeleton to graph and analyze the graph
    
    Inputs: 
    
        image_skeleton: skeleton image accquire from image binary mask

    Returns:
    
        branch_data:structure of branch info
        ###
        # ['skeleton-id', 'node-id-src', 'node-id-dst', 'branch-distance', 
        #'branch-type', 'mean-pixel-value', 'stdev-pixel-value', 
        #'image-coord-src-0', 'image-coord-src-1', 'image-coord-dst-0', 'image-coord-dst-1', 
        #'coord-src-0', 'coord-src-1', 'coord-dst-0', 'coord-dst-1', 'euclidean-distance']
        ###
        
        n_branch: number of brunches 
        
        sub_branch_branch_distance_cleaned: list of individual branch length 
        
    """
        

    #get brach data
    branch_data = summarize(Skeleton(image_skeleton))
    #print(branch_data)
    
    #select end branch
    sub_branch = branch_data.loc[branch_data['branch-type'] == 1]
    
    sub_branch_cleaned = sub_branch
    
    
    sub_branch_branch_distance = sub_branch["branch-distance"].tolist()
    
    sub_branch_branch_distance_cleaned = sub_branch_branch_distance
    
    '''
    # remove outliers in branch distance 
    outlier_list = outlier_doubleMAD(sub_branch_branch_distance, thresh = 3.5)
    
    indices = [i for i, x in enumerate(outlier_list) if x]
    
    sub_branch_cleaned = sub_branch.drop(sub_branch.index[indices])
    
    sub_branch_branch_distance_cleaned = [i for j, i in enumerate(sub_branch_branch_distance) if j not in indices]
    '''
    
    print("Branch graph info : {0}\n".format(sub_branch_cleaned))
    

    branch_type_list = sub_branch_cleaned["branch-type"].tolist()

    n_branch = branch_type_list.count(1)

    
    return branch_data, n_branch, sub_branch_branch_distance_cleaned



#apply CLAHE (Contrast Limited Adaptive Histogram Equalization) to perfrom image enhancement
def image_enhance(img):

    # CLAHE (Contrast Limited Adaptive Histogram Equalization)
    clahe = cv2.createCLAHE(clipLimit=3., tileGridSize=(8,8))

    # convert from BGR to LAB color space
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)  
    
    # split on 3 different channels
    l, a, b = cv2.split(lab)  

    # apply CLAHE to the L-channel
    l2 = clahe.apply(l)  

    # merge channels
    lab = cv2.merge((l2,a,b))  
    
    # convert from LAB to BGR
    img_enhance = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)  
    
    return img_enhance




def adjust_gamma(image, gamma):

    """Adjust the gamma value to increase the brightness of image
    
    Inputs: 
    
        image: image 
        
        gamma: gamma value used to adjust

    Returns:
    
        cv2.LUT(image, table): adjusted image with gamma correction
        
        
    """
    
    # build a lookup table mapping the pixel values [0, 255] to
    # their adjusted gamma values
    invGamma = 1.0 / gamma
    table = np.array([((i / 255.0) ** invGamma) * 255
        for i in np.arange(0, 256)]).astype("uint8")
 
    # apply gamma correction using the lookup table
    return cv2.LUT(image, table)
    
    

def extract_traits(image_file):

    """compute all the traits based on input image
    
    Inputs: 
    
        image file: full path and file name

    Returns:
        image_file_name: file name
        
        tag_info: Barcode information
        
        tassel_area: area occupied by the tassel in the image
        
        tassel_area_ratio: The ratio between tassel area and its convex hull area
        
        cnt_width, cnt_height: width and height of the tassel
        
        n_branch: number of branches in the tassel
        
        avg_branch_length: average length of the branches
        
        branch_length: list of all branch length
    """
    
    # extarct path and name of the image file
    abs_path = os.path.abspath(image_file)
    
    filename, file_extension = os.path.splitext(abs_path)
    
    # extract the base name 
    base_name = os.path.splitext(os.path.basename(filename))[0]

    # get the file size
    file_size = os.path.getsize(image_file)/MBFACTOR
    
    # get the image file name
    image_file_name = Path(image_file).name


    print("Exacting traits for image : {0}\n".format(str(image_file_name)))
     
    # create result folder
    if (args['result']):
        save_path = args['result']
    else:
        mkpath = os.path.dirname(abs_path) +'/' + base_name
        mkdir(mkpath)
        save_path = mkpath + '/'
        

    print ("results_folder:" + save_path +'\n')
    

    # initialize all the traits output 
    tag_info = tassel_area = tassel_area_ratio = cnt_width = cnt_height = 0
    
    n_branch = avg_branch_length = coins_width_avg = 0
    
     
    if (file_size > 5.0):
        print("It will take some time due to larger file size {0} MB\n".format(str(int(file_size))))
    else:
        print("Perform plant object segmentation using automatic color clustering method...\n")
    
    # load the input image 
    image = cv2.imread(image_file)
    
    #make backup image
    orig = image.copy()
    
    # get the dimension of the image
    img_height, img_width, img_channels = orig.shape
    
    #source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
    
    # parse input arguments
    args_colorspace = args['color_space']
    args_channels = args['channels']
    args_num_clusters = args['num_clusters']
    
    
    
    #color clustering based object segmentation to accquire external contours
    image_mask = color_cluster_seg(image.copy(), args_colorspace, args_channels, args_num_clusters)
    
    
    # save segmentation result
    result_file = (save_path + base_name + '_image_mask' + file_extension)
    cv2.imwrite(result_file, image_mask)

    ####################################################################
    # check segmentation mask
    # if mask was empty or only coin and barcode was detected, assign 0 values 
    
    #find contours and get the external one
    (contours, hier) = cv2.findContours(image_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) > 0:
            
        # sort the contours based on area from largest to smallest
        contours_sorted = sorted(contours, key = cv2.contourArea, reverse = True)
        
        max_area = cv2.contourArea(contours_sorted[0])
    else:
        
        max_area = 0
        
    #print("max_area = {}\n".format(max_area))
    ########################################################################
    
    # make sure tassel object was detected
    if (cv2.countNonZero(image_mask) == 0) or (max_area < min_size):
        
        print("Mask image is empty...\n")
    else:
        
        print("Mask segmentation finished successfully!\n")

        
        ################################################################################################################################
        #compute external traits
        (trait_img, tassel_area, cnt_width, cnt_height, tassel_area_ratio) = comp_external_contour(image.copy(), image_mask)
        
        # save segmentation result
        result_file = (save_path + base_name + '_excontour' + file_extension)
        #print(filename)
        cv2.imwrite(result_file, trait_img)
        
        
        #get the medial axis of the contour
        image_skeleton, skeleton = skeleton_bw(image_mask)

        # save skeleton result
        result_file = (save_path + base_name + '_skeleton' + file_extension) 
        cv2.imwrite(result_file, img_as_ubyte(image_skeleton))
        
        
        (branch_data, n_branch, branch_length) = skeleton_graph(image_skeleton)
        
        avg_branch_length = int(sum(branch_length)/len(branch_length))


        #img_hist = branch_data.hist(column = 'branch-distance', by = 'branch-type', bins = 100)
        #result_file = (save_path + base_name + '_hist' + file_extension)
        #plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        #plt.close()
        '''
        fig = plt.plot()
        source_image = cv2.cvtColor(orig, cv2.COLOR_BGR2RGB)
        #img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-distance', skeleton_colormap = 'hsv')
        img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-type', skeleton_colormap = 'hsv')
        
        # save skeleton result
        result_file = (save_path + base_name + '_skeleton_graph' + file_extension) 
        plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0, dpi = 600)
        plt.close()
        '''
     
        #print("[INFO] {} branch end points found\n".format(n_branch))
        
        #print("[INFO] {} branch length\n".format(branch_length))
        
    
    
    ###################################################################################################
    # detect coin and barcode uisng template mathcing method
    
    # define right bottom area for coin detection
    x = int(img_width*0.5)
    y = int(img_height*0.5)
    w = int(img_width*0.5)
    h = int(img_height*0.5)
    
    roi_image = region_extracted(orig, x, y, w, h)
    
    # apply gamma correction for image region with coin
    gamma = 1.5
    gamma = gamma if gamma > 0 else 0.1
    enhanced_region = adjust_gamma(roi_image.copy(), gamma=gamma)
    
    (circles, circle_detection_img, diameter_circle) = circle_detection(enhanced_region) 
    
    #print("coin_size = {}".format(coin_size))
    
    pixel_cm_ratio = diameter_circle/coin_size
    
    coins_width_avg = diameter_circle
    
    # save result
    result_file = (save_path + base_name + '_coin_circle' + file_extension)
    cv2.imwrite(result_file, circle_detection_img)
    
    
    print("The width of coin in the marker image is {:.0f} × {:.0f} pixels\n".format(diameter_circle, diameter_circle))
    
    
    
    # define left bottom area for barcode detection
    x = 0
    y = int(img_height*0.5)
    w = int(img_width*0.5)
    h = int(img_height*0.5)
    
    roi_image = region_extracted(orig, x, y, w, h)
    
    # method for template matching 
    methods = ['cv.TM_CCOEFF', 'cv.TM_CCOEFF_NORMED', 'cv.TM_CCORR',
        'cv.TM_CCORR_NORMED', 'cv.TM_SQDIFF', 'cv.TM_SQDIFF_NORMED']
    
    # apply gamma correction for image region with coin
    gamma = 1.5
    gamma = gamma if gamma > 0 else 0.1
    enhanced_region = adjust_gamma(roi_image.copy(), gamma=gamma)
    
    # detect the barcode object based on template image
    (marker_barcode_img, thresh_barcode, barcode_width_contour, barcode_width_circle) = marker_detect(enhanced_region, tp_barcode, methods[0], 0.8)
    
    # save result
    result_file = (save_path + base_name + '_barcode' + file_extension)
    cv2.imwrite(result_file, marker_barcode_img)
    
    # save result
    #result_file = (save_path + base_name + '_barcode_mask' + file_extension)
    #cv2.imwrite(result_file, thresh_barcode)
    
    # parse barcode image using pylibdmtx lib
    tag_info = barcode_detect(marker_barcode_img)

    ####################################################
    
    return image_file_name, tag_info, tassel_area, tassel_area_ratio, cnt_width, cnt_height, n_branch, avg_branch_length, coins_width_avg, coin_size, pixel_cm_ratio
    





if __name__ == '__main__':
    
    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required = True,    help = "path to image file")
    ap.add_argument("-ft", "--filetype", required = True,    help = "Image filetype")
    ap.add_argument('-mk', '--marker', required = False,  default ='/marker_template/coin.png',  help = "Marker file name")
    ap.add_argument('-bc', '--barcode', required = False,  default ='/marker_template/barcode.png',  help = "Barcode file name")
    ap.add_argument("-r", "--result", required = False,    help="result path")
    ap.add_argument('-s', '--color-space', type = str, required = False, default ='lab', help='Color space to use: BGR, HSV, Lab (default), YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type = str, required = False, default='2', help='Channel indices to use for clustering, where 0 is the first channel,' 
                                                                       + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" ' 
                                                                       + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type = int, required = False, default = 2,  help = 'Number of clusters for K-means clustering (default 2, min 2).')
    ap.add_argument('-min', '--min_size', type = int, required = False, default = 35000,  help = 'min size of object to be segmented.')
    ap.add_argument('-cs', '--coin_size', type = int, required = False, default = 2.7,  help = 'coin size in cm')
    
    args = vars(ap.parse_args())
    
    
    # parse input arguments
    file_path = args["path"]
    ext = args['filetype']
    
    coin_path = args["marker"]
    barcode_path = args["barcode"]
    
    min_size = args['min_size']
    coin_size = args['coin_size']
    
    
    # path of the marker (coin), default path will be '/marker_template/marker.png' and '/marker_template/barcode.png'
    # can be changed based on requirement
    global  tp_coin, tp_barcode

    
    #setup marker path to load template
    template_path = file_path + coin_path
    barcode_path = file_path + barcode_path
    
    try:
        # check to see if file is readable
        with open(template_path) as tempFile:

            # Read the template 
            tp_coin = cv2.imread(template_path, 0)
            print("Template loaded successfully...")
            
    except IOError as err:
        
        print("Error reading the Template file {0}: {1}".format(template_path, err))
        exit(0)

    
    try:
        # check to see if file is readable
        with open(barcode_path) as tempFile:

            # Read the template 
            tp_barcode = cv2.imread(barcode_path, 0)
            print("Barcode loaded successfully...\n")
            
    except IOError as err:
        
        print("Error reading the Barcode file {0}: {1}".format(barcode_path, err))
        exit(0)
    
    

    #accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype
    
    #accquire image file list
    imgList = sorted(glob.glob(image_file_path))
    
    imgList = natsort.natsorted(imgList,reverse = False)

    #print((imgList))
    
    n_images = len(imgList)
    
    result_list = []
    
    result_list_cm = []
    
    
    #loop execute to get all traits
    for image in imgList:
        
        (filename, tag_info, tassel_area, tassel_area_ratio, avg_width, avg_height, n_branch, avg_branch_length, coins_width_avg, coin_size, pixel_cm_ratio) = extract_traits(image)

        result_list.append([filename, tag_info, tassel_area, tassel_area_ratio, avg_width, avg_height, n_branch, avg_branch_length, coins_width_avg, coin_size, pixel_cm_ratio])
        
        result_list_cm.append([filename,tag_info,tassel_area/pow(pixel_cm_ratio,2),tassel_area_ratio,avg_width/pixel_cm_ratio,avg_height/pixel_cm_ratio,n_branch,avg_branch_length/pixel_cm_ratio,coins_width_avg/pixel_cm_ratio,coin_size,pixel_cm_ratio])
        
        
        #for i in range(len(branch_length)):
            
            #result_list_cm.append([filename, str(i).zfill(2), branch_length[i]])

    '''
    ###########################################################
    #parallel processing module
    
    # get cpu number for parallel processing
    agents = psutil.cpu_count() - 2 
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result = pool.map(extract_traits, imgList)
        pool.terminate()
        
    # unwarp all the computed trait values
    filename = list(zip(*result))[0]
    tag_info = list(zip(*result))[1]
    tassel_area = list(zip(*result))[2]
    tassel_area_ratio = list(zip(*result))[3]
    avg_width = list(zip(*result))[4]
    avg_height = list(zip(*result))[5]
    n_branch = list(zip(*result))[6]
    avg_branch_length = list(zip(*result))[7]
    coins_width_avg = list(zip(*result))[8]
    coin_size = list(zip(*result))[9]
    pixel_cm_ratio = list(zip(*result))[10]

    # create result list
    for i, (v0,v1,v2,v3,v4,v5,v6,v7,v8,v9,v10) in enumerate(zip(filename, tag_info, tassel_area, tassel_area_ratio, avg_width, avg_height, n_branch, avg_branch_length, coins_width_avg, coin_size, pixel_cm_ratio)):

        result_list.append([v0,v1,v2,v3,v4,v5,v6,v7,v8,v9,v10])
        
        result_list_cm.append([v0,v1,v2/pow(v10,2),v3,v4/v10,v5/v10,v6,v7/v10,v8/v10,v9,v10])
    '''

    #################################################################################
    #print out result on screen output as table
    
    print("Summary: {0} plant images were processed...\n".format(n_images))
    
    #output in command window in a sum table
 
    #table = tabulate(result_list, headers = ['filename', 'mazie_ear_area', 'kernel_area_ratio', 'max_width', 'max_height' ,'cluster 1', 'cluster 2', 'cluster 3', 'cluster 4', 'cluster 1 hex value', 'cluster 2 hex value', 'cluster 3 hex value', 'cluster 4 hex value'], tablefmt = 'orgtbl')
    table = tabulate(result_list, headers = ['filename', 'tag_info', 'tassel_area', 'tassel_area_ratio', 'avg_width', 'avg_height', 'number_branches', 'average branch_length', 'coins_width_avg', 'coin_size', 'pixel_cm_ratio'], tablefmt = 'orgtbl')
    
    print(table + "\n")
    
    

    #####################################################################
    # save computation traits results as excel file
    
    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')
        trait_file_csv = (args['result'] + 'trait.csv')
    else:
        trait_file = (file_path + 'trait.xlsx')
        trait_file_csv = (file_path + 'trait.csv')
    

    if os.path.exists(trait_file):
        
        # update values
        #Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        #Get the current Active Sheet
        sheet_pixel = wb['trait_pixel']

        sheet_pixel.delete_rows(2, sheet_pixel.max_row - 1) # for entire sheet

        #Get the current Active Sheet
        sheet_cm = wb['trait_cm']

        sheet_cm.delete_rows(2, sheet_cm.max_row - 1) # for entire sheet
        

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        
        sheet_pixel = wb.active
        sheet_pixel.title = "trait_pixel"
        
        sheet_cm = wb.create_sheet("Sheet2")
        sheet_cm.title = "trait_cm"

    sheet_pixel.cell(row = 1, column = 1).value = 'filename'
    sheet_pixel.cell(row = 1, column = 2).value = 'tag info'
    sheet_pixel.cell(row = 1, column = 3).value = 'tassel area'
    sheet_pixel.cell(row = 1, column = 4).value = 'tassel area ratio'
    sheet_pixel.cell(row = 1, column = 5).value = 'average width'
    sheet_pixel.cell(row = 1, column = 6).value = 'average height'
    sheet_pixel.cell(row = 1, column = 7).value = 'number of branches'
    sheet_pixel.cell(row = 1, column = 8).value = 'average branch length'
    sheet_pixel.cell(row = 1, column = 9).value = 'average coins width'
    sheet_pixel.cell(row = 1, column = 10).value = 'coin size'
    sheet_pixel.cell(row = 1, column = 11).value = 'pixel/cm ratio'

    sheet_cm.cell(row = 1, column = 1).value = 'filename'
    sheet_cm.cell(row = 1, column = 2).value = 'tag info'
    sheet_cm.cell(row = 1, column = 3).value = 'tassel area'
    sheet_cm.cell(row = 1, column = 4).value = 'tassel area ratio'
    sheet_cm.cell(row = 1, column = 5).value = 'average width'
    sheet_cm.cell(row = 1, column = 6).value = 'average height'
    sheet_cm.cell(row = 1, column = 7).value = 'number of branches'
    sheet_cm.cell(row = 1, column = 8).value = 'average branch length'
    sheet_cm.cell(row = 1, column = 9).value = 'average coins width'
    sheet_cm.cell(row = 1, column = 10).value = 'coin size'
    sheet_cm.cell(row = 1, column = 11).value = 'pixel/cm ratio'



    # write traits values
    for row in result_list:
        sheet_pixel.append(row)
    
    for row in result_list_cm:
        sheet_cm.append(row)
        
   
    
    #save computation traits results as xlsx format excel file
    wb.save(trait_file)
    
    
    


    

    
